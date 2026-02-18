import re
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from pandas import DataFrame as _DataFrame
import os
import io
import zipfile
import unicodedata
from io import BytesIO
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title='IA4Doc ‚Äì KPI & Analyse Fiches', layout='wide')

def verdict_to_score(verdict: object) -> float:
    """Mappe strictement le verdict vers un score : bon=1, partiel=0.5, mauvais=0."""
    if not isinstance(verdict, str):
        return np.nan
    
    v = verdict.strip() # On garde juste le retrait des espaces inutiles
    
    if v == "Bon":
        return 1.0
    if v == "Partiellement bon":
        return 0.5
    if v == "Mauvais":
        return 0.0
        
    return np.nan # Si c'est vide ou √©crit autrement

# ---------------------------------------------------------------------

def extract_doc_id(ref: str) -> str:
    """Extrait un identifiant de document depuis la r√©f√©rence (nom de fichier).
    Heuristique : cherche d'abord un motif explicite (doc/document/coedm/ref + chiffres),
    sinon prend la premi√®re s√©quence de >=4 chiffres, sinon retourne la r√©f√©rence enti√®re.
    """
    if not isinstance(ref, str):
        return ""
    s = ref.strip()
    m = re.search(r"(?:doc|document|coedm|ref)[-_ ]*(\d+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4,})", s)
    if m:
        return m.group(1)
    return s



# =====================================================================
# PARSE FICHE
# =====================================================================

#permet de lire les cases des fiches de test

def parse_fiche(file):
    filename = file.name

    if filename.startswith("~$"):
        return pd.DataFrame()  # fichier temporaire Excel => on ignore (j'ai d√©ja rencontr√© le cas, d'o√π cette s√©curit√©)

    # Exemples de noms :
    # - fs60-IA-v00-TUF-FFP_PV_...-Fiche-v01-CM.xlsx
    # - fs1-IA-v00-NR-...-Fiche-v01-CM.xlsx (fs 1 √† 3 chiffres)
    filename_clean = str(filename).strip()

    # --- 1. Analyse du NOM du fichier ---
    # V√©rifie si le nom respecte le format strict fsXX-IA-vXX..
    pat = r"^(fs\d{1,3})-IA-v\d+-(TUF|NR)-(.+)-Fiche-v\d+-[A-Za-z]+\.xlsx$" # si le nom des fiches change, c'est ici qu'il faut faire les modifs !
    m = re.match(pat, filename_clean, flags=re.IGNORECASE)
    if not m:
        # Si le nom est mauvais, le script s'arr√™te ici pour ce fichier
        raise ValueError(
            "Nom de fichier invalide. Format attendu : "
            "fs<1-3 chiffres>-IA-vXX-(TUF|NR)-<ref>-Fiche-vXX-<initiales>.xlsx"
        )

    fs_id = m.group(1).lower()
    fiche_type = m.group(2).upper()
    ref_from_filename = m.group(3)
    doc_id = extract_doc_id(ref_from_filename)

    # apr√®s upload des fiches de test, l'outil va lire la feuille "Template Fiche de Test" <- si le nom de la feuille excel change c'est ici qu'il faut modifier 
    df = pd.read_excel(file, sheet_name="Template Fiche de Test", header=None)

    # -----------------------------------------------------------------
    # D√©tection du d√©calage : ligne 18 contient-elle "*Searchable:" ? (apr√®s la nouvelle version du template des fiches de test)
    # -----------------------------------------------------------------
    has_searchable_row = False
    try:
        row18 = df.iloc[17]  # ligne 18 (index 17)
        for val in row18:
            if isinstance(val, str) and "Searchable" in val:
                has_searchable_row = True
                break
    except Exception:
        pass

    offset = 0 if has_searchable_row else -1

    def r(base_row):
        # applique le d√©calage seulement apr√®s la ligne 18
        if base_row > 17:
            return base_row + offset
        return base_row

    # ===========================
    # Champs "avant 18" R√©cup√©ration des infos g√©n√©rales (En-t√™te de fiche)
    # ===========================
    type_test = df.iloc[4, 1] if pd.notna(df.iloc[4, 1]) else None
    type_doc = df.iloc[14, 1] if pd.notna(df.iloc[14, 1]) else None
    ref_coedm = ref_from_filename

    # Label fonctionnalit√© (ligne 24 ‚Üí index 23, apr√®s 18 ‚Üí offset)
    label_fonctionnalite = None
    try:
        val_lbl = df.iloc[r(23), 1]
        if pd.notna(val_lbl):
            label_fonctionnalite = val_lbl
    except Exception:
        pass

    if label_fonctionnalite is None or (
        isinstance(label_fonctionnalite, float) and pd.isna(label_fonctionnalite)
    ):
        label_fonctionnalite = fs_id

    # Date (B10 ‚Üí ligne 10, index 9) = avant 18
    date_test = None
    try:
        val_date = df.iloc[9, 1]
        if not pd.isna(val_date):
            date_test = pd.to_datetime(val_date, errors="coerce")
    except Exception:
        pass

    # Site (B9 ‚Üí ligne 9, index 8, colonne 1)
    site = None
    try:
        val_site = df.iloc[8, 1]  # B9
        if not pd.isna(val_site):
            site = str(val_site).strip()
            if site == "LTA":
                site = "STMA"
    except Exception:
        pass

    # Code modification (B22)
    code_modif = None
    try:
        val_modif = df.iloc[21, 1]  # B22
        if isinstance(val_modif, str) and val_modif.strip():
            code_modif = val_modif.strip()
    except Exception:
        pass
    

    # Classe documentaire (D15 ‚Üí ligne 15, index 14) = avant 18
    classe_documentaire = None
    try:
        val_cls = df.iloc[14, 3]
        if pd.isna(val_cls):
            classe_documentaire = None
        elif isinstance(val_cls, str):
            v = val_cls.strip()
            if not v:
                classe_documentaire = None
            elif v.lower() == "non":
                classe_documentaire = "Non Searchable"
            elif v.lower() == "oui":
                classe_documentaire = "Searchable"
            else:
                classe_documentaire = v
        else:
            classe_documentaire = val_cls
    except Exception:
        pass

    # Nombre de pages total (ligne 17 ‚Üí index 16)
    try:
        nb_pages_total = int(df.iloc[16, 1])
    except Exception:
        nb_pages_total = None

    # Commentaire additionnel (ligne 46 ‚Üí index 45, apr√®s 18 ‚Üí offset)
    commentaire_add = None
    try:
        comment_row = df.iloc[r(45)]
        for val in comment_row:
            if isinstance(val, str) and val.strip():
                commentaire_add = val.strip()
                break
    except Exception:
        pass

    # Verdict doc : lignes 6 √† 12 (index 5..11)
    verdict_doc = None
    try:
        for rr in range(5, 12):
            val = df.iloc[rr, 2]
            if isinstance(val, str) and val.strip():
                verdict_doc = val.strip()
                break
    except Exception:
        pass

    verdict_score = verdict_to_score(verdict_doc)

    # Nom testeur : lignes 6 et 7 (index 5 & 6)
    nom_testeur = None
    try:
        val7 = df.iloc[6, 1]
        val6 = df.iloc[5, 1]
        out = []
        if isinstance(val7, str) and val7.strip():
            out.append(val7.strip())
        if isinstance(val6, str) and val6.strip():
            out.append(val6.strip())
        if out:
            nom_testeur = " ".join(out)
    except Exception:
        pass

    # Fonctionnalit√© : 25A / 26A avec offset
    fonctionnalite = None
    try:
        cell_26A = df.iloc[r(25), 0]  # ligne 26 (index 25)
        cell_25A = df.iloc[r(24), 0]  # ligne 25 (index 24)
        sentinel = "[NR] Nombre de test de r√©p√©tabilit√© requis"

        if isinstance(cell_26A, str) and cell_26A.strip() and cell_26A.strip() != sentinel:
            fonctionnalite = cell_26A.strip()
        elif isinstance(cell_25A, str) and cell_25A.strip():
            fonctionnalite = cell_25A.strip()
    except Exception:
        pass

    # Noms des tests (ligne 31 ‚Üí index 30, apr√®s 18 ‚Üí offset)
    try:
        tests = list(df.iloc[r(30), 2:6].dropna())
    except Exception:
        tests = []

    records = []
    # Pour chaque test trouv√©...
    for k in range(len(tests)):
        col = 2 + k # D√©calage de colonnes (C=2, D=3...)

        # lignes 32..37 ‚Üí index 31..36 ‚Üí apr√®s 18 ‚Üí offset
        # On r√©cup√®re les lignes de r√©sultats (Temps, FN, FP...)
        row_h = r(31)
        row_m = r(32)
        row_j = r(33)
        row_fn = r(34)
        row_fp = r(35)
        row_inc = r(36)

        raw_vals = [
            df.iloc[row_h, col],   # humain
            df.iloc[row_m, col],   # machine
            df.iloc[row_j, col],   # justes
            df.iloc[row_fn, col],  # fn
            df.iloc[row_fp, col],  # fp
            df.iloc[row_inc, col], # incertaines
        ]

        # Test totalement vide ‚Üí ignor√©
        if all(pd.isna(v) for v in raw_vals):
            continue

        def to_float_or_none(v):
            try:
                return float(v) if not pd.isna(v) else None
            except Exception:
                return None

        humain = to_float_or_none(raw_vals[0])
        machine = to_float_or_none(raw_vals[1])
        justes = to_float_or_none(raw_vals[2])
        fn = to_float_or_none(raw_vals[3])
        fp = to_float_or_none(raw_vals[4])
        inc = to_float_or_none(raw_vals[5])

        # Cr√©ation de l'objet r√©sultat
        rec = {
            "fs_id": fs_id,
            "ref_coedm": ref_coedm,
            "doc_id": doc_id,
            "fiche_type": fiche_type,
            "fiche_name": filename,
            "date_test": date_test,
            "type_test": type_test,
            "type_document": type_doc,
            "classe_documentaire": classe_documentaire,
            "site": site,
            "label_fonctionnalite": label_fonctionnalite,
            "code_modif": code_modif,
            "nb_pages_total": nb_pages_total,
            "commentaire_additionnel": commentaire_add,
            "verdict_doc": verdict_doc,
            "verdict_score": verdict_score,
            "nom_testeur": nom_testeur,
            "fonctionnalite": fonctionnalite,
            "test_label": f"Test {k+1}",
            "temps_humain_s": humain,
            "temps_machine_s": machine,
            "nb_pages_justes": justes,
            "fn": fn,
            "fp": fp,
            "incertaines": inc,
        }
        # Calcul des pourcentages
        if nb_pages_total and nb_pages_total != 0:
            rec["taux_justes"] = (justes / nb_pages_total * 100) if justes is not None else None
            rec["taux_fn"] = (fn / nb_pages_total * 100) if fn is not None else None
            rec["taux_fp"] = (fp / nb_pages_total * 100) if fp is not None else None
            rec["taux_incertaines"] = (inc / nb_pages_total * 100) if inc is not None else None
        else:
            rec["taux_justes"] = None
            rec["taux_fn"] = None
            rec["taux_fp"] = None
            rec["taux_incertaines"] = None


        if humain is not None and machine is not None:
            rec["gain_temps_s"] = humain - machine
        else:
            rec["gain_temps_s"] = None

        records.append(rec)

    return pd.DataFrame(records)


# =====================================================================
# INTERFACE STREAMLIT
# =====================================================================

st.title("Analyse automatique des fiches IA4Doc")

# Zone de glisser-d√©poser des fichiers
uploaded_files = st.file_uploader(
    "Fichiers Excel ou ZIP",
    type=["xlsx", "zip"],
    accept_multiple_files=True
,
    key="fiches_top"
)

if uploaded_files:
    st.session_state["fiches_uploaded_files"] = uploaded_files
else:
    st.session_state.setdefault("fiches_uploaded_files", [])



def iter_excel_files(uploaded_files):
    """
    G√©n√®re des objets fichier Excel (avec un attribut .name)
    √† partir de ce que l'utilisateur a upload√© :
    - .xlsx directs
    - .zip contenant des .xlsx
    """
    for up in uploaded_files:
        fname = up.name.lower()

        # 1) Cas ZIP
        if fname.endswith(".zip"):
            try:
                # Remet le curseur au d√©but (Streamlit peut relire le m√™me objet)
                try:
                    up.seek(0)
                except Exception:
                    pass

                zbytes = up.read()
                with zipfile.ZipFile(io.BytesIO(zbytes)) as zf:
                    for member in zf.namelist():
                        if member.endswith("/") or (not member.lower().endswith(".xlsx")):
                            continue
                        data = zf.read(member)
                        bio = io.BytesIO(data)
                        bio.name = os.path.basename(member)
                        try:
                            bio.seek(0)
                        except Exception:
                            pass
                        yield bio
            except Exception as e:
                st.error(f"Erreur en lisant le zip {up.name} : {e}")# 2) Cas Excel direct
        elif fname.endswith(".xlsx"):
            yield up


# =============================================================
# KPI (Tableaux + Graphiques) ‚Äî AU D√âBUT
# =============================================================

from datetime import datetime, time as _time, date as _date
from openpyxl.utils import get_column_letter

# --- CDC parsing rules (prompt KPI)
_IB_SET = {"IBT", "IC"}
_PCM_SET = {"STMA", "JMT", "FLC"}

def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")

def _to_excel_datetime(d: _date) -> datetime:
    return datetime.combine(d, _time.min)

def _excel_col_letters(start: str, end: str) -> list[str]:
    def col_to_num(c):
        n = 0
        for ch in c:
            n = n*26 + (ord(ch.upper()) - 64)
        return n
    def num_to_col(n):
        s = ""
        while n:
            n, r = divmod(n-1, 26)
            s = chr(65+r) + s
        return s
    a, b = col_to_num(start), col_to_num(end)
    return [num_to_col(i) for i in range(a, b+1)]

_COMMON_MARK_COLS = _excel_col_letters("AK", "AR")

def parse_cdc_functions(cdc_bytes: bytes) -> pd.DataFrame:
    """Lit le CDC (feuille 1-Fonctionnalit√©s) et renvoie : ID fonctionnalit√©, Cat√©gorie, √âtat"""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(cdc_bytes), data_only=True)
    if "1-Fonctionnalit√©s" not in wb.sheetnames:
        raise ValueError("Feuille '1-Fonctionnalit√©s' introuvable dans le CDC.")
    ws = wb["1-Fonctionnalit√©s"]

    rows = []
    for r in range(5, ws.max_row + 1):
        func_id = ws[f"B{r}"].value
        if _is_blank(func_id):
            continue

        flag_a = ws[f"A{r}"].value
        origin = ws[f"C{r}"].value
        state = ws[f"AI{r}"].value

        has_mark = False
        for col in _COMMON_MARK_COLS:
            v = ws[f"{col}{r}"].value
            if not _is_blank(v):
                has_mark = True
                break

        rows.append({
            "row": r,
            "flag_a": flag_a,
            "func_id": str(func_id).strip(),
            "origin": str(origin).strip() if not _is_blank(origin) else "",
            "state": str(state).strip() if not _is_blank(state) else "",
            "has_mark": has_mark,
        })

    raw = pd.DataFrame(rows)
    if raw.empty:
        return raw

    def flag_is_one(x) -> bool:
        if x is None:
            return False
        s = str(x).strip().lower()
        return s == "1" or s == "oui"

    out_rows = []
    for func_id, g in raw.groupby("func_id", sort=False):
        g = g.copy()
        base = g[g["flag_a"].apply(flag_is_one)]
        base_row = base.iloc[0] if len(base) else g.iloc[0]

        base_origin = (base_row["origin"] or "").strip()
        base_state = (base_row["state"] or "").strip()

        is_common = (base_origin.lower() == "ib et pcm") and bool(g["has_mark"].any())
        if is_common:
            marked = g[g["has_mark"] == True]
            marked_origins = set([str(o).strip() for o in marked["origin"].tolist() if not _is_blank(o)])
            has_ib = any(o in _IB_SET for o in marked_origins)
            has_pcm = any(o in _PCM_SET for o in marked_origins)

            if has_ib and has_pcm:
                category = "Commun universel"
            elif has_ib:
                category = "Commun sp√©cifique IB"
            elif has_pcm:
                category = "Commun sp√©cifique PCM"
            else:
                category = "Commun universel"
        else:
            bo = base_origin.strip()
            if bo == "PCM":
                category = "PCM"
            elif bo == "IB":
                category = "IB"
            elif bo.lower() == "ib et pcm":
                non_flag = g[~g["flag_a"].apply(flag_is_one)]
                origins = set([str(o).strip() for o in non_flag["origin"].tolist() if not _is_blank(o)])
                if any(o in _IB_SET for o in origins):
                    category = "IB"
                elif any(o in _PCM_SET for o in origins):
                    category = "PCM"
                else:
                    category = "IB"
            else:
                category = bo if bo else "Autre"

        out_rows.append({
            "ID fonctionnalit√©": func_id,
            "Cat√©gorie": category,
            "√âtat": base_state,
        })

    return pd.DataFrame(out_rows)

# --- Historique tableaux_hebdo.xlsx (template)
_T1_TITLE = "Nombre de fonctionnalit√©s en test"
_T3_TITLE = "Nombre de Fiche de Test remplies"
_T2_TITLE = "Etat des fonctionnalit√©s"

# dans votre template:
# - Tableau 1 : titre en B3, dates en C3.., labels en B4..B8
# - Tableau 3 : titre en B12, dates en C12.., labels en B13.. (types doc)
# - Tableau 2 : dates en C20.. (B20 vide), labels en B21..B27 dont Total

def _find_row_by_title(ws, title: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip() == title:
            return r
    return None


def _strip_accents(s: str) -> str:
    return ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')

def _find_row_by_title_fuzzy(ws, title: str) -> int | None:
    """Find row where column B matches title (accent/case-insensitive)."""
    target = _strip_accents(title).strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str):
            vv = _strip_accents(v).strip().lower()
            if vv == target:
                return r
    return None

def _detect_table_block_height(ws, labels_start_row: int) -> int:
    """Detect number of label rows until blank (includes 'Total' row if present)."""
    r = labels_start_row
    seen = 0
    while r <= ws.max_row + 50:
        v = ws.cell(r, 2).value
        if _is_blank(v):
            break
        seen += 1
        if isinstance(v, str) and _strip_accents(v).strip().lower() == 'total':
            break
        r += 1
    return seen

def _move_values(ws, src_row: int, dst_row: int, n_rows: int, n_cols: int) -> None:
    """Move (cut/paste) values in a rectangular range."""
    if n_rows <= 0 or n_cols <= 0 or src_row == dst_row:
        return
    tmp = []
    for r_off in range(n_rows):
        row_vals = []
        for c in range(1, n_cols + 1):
            row_vals.append(ws.cell(src_row + r_off, c).value)
        tmp.append(row_vals)
    # clear src
    for r_off in range(n_rows):
        for c in range(1, n_cols + 1):
            ws.cell(src_row + r_off, c).value = None
    # paste
    for r_off in range(n_rows):
        for c in range(1, n_cols + 1):
            ws.cell(dst_row + r_off, c).value = tmp[r_off][c-1]

def _normalize_table2_location(ws) -> tuple[int, int]:
    """Ensure Tableau 2 is anchored at row 20 (title in B20). Returns (title_row, labels_start_row)."""
    TARGET_ROW = 20
    found_row = _find_row_by_title_fuzzy(ws, _T2_TITLE)
    title_row = found_row or TARGET_ROW
    labels_start = title_row + 1
    # If table exists elsewhere and B20 is blank, move it to row 20 to keep a single source of truth.
    if found_row and found_row != TARGET_ROW and _is_blank(ws.cell(TARGET_ROW, 2).value):
        height = _detect_table_block_height(ws, found_row + 1)
        dates = _read_date_headers(ws, found_row, start_col=3)
        n_rows = 1 + height
        n_cols = 2 + max(1, len(dates))
        _move_values(ws, found_row, TARGET_ROW, n_rows, n_cols)
        title_row = TARGET_ROW
        labels_start = TARGET_ROW + 1
    # Ensure title exists at anchor
    if _is_blank(ws.cell(TARGET_ROW, 2).value):
        ws.cell(TARGET_ROW, 2).value = _T2_TITLE
    return TARGET_ROW, TARGET_ROW + 1
def _read_date_headers(ws, row: int, start_col: int = 3) -> list[_date]:
    out = []
    c = start_col
    while True:
        v = ws.cell(row, c).value
        if v is None:
            break
        if isinstance(v, datetime):
            out.append(v.date())
        elif isinstance(v, _date):
            out.append(v)
        else:
            try:
                out.append(pd.to_datetime(v).date())
            except Exception:
                break
        c += 1
    return out

def _find_or_append_date_col(ws, row: int, snap: _date, start_col: int = 3) -> int:
    dates = _read_date_headers(ws, row, start_col=start_col)
    if snap in dates:
        return start_col + dates.index(snap)
    col = start_col + len(dates)
    ws.cell(row, col).value = _to_excel_datetime(snap)
    return col

def _read_block(ws, title_row: int, labels_start_row: int, stop_on_blank: bool = True) -> pd.DataFrame:
    """Retourne df index=labels(col B), columns=dates (date headers row=title_row)"""
    dates = _read_date_headers(ws, title_row, start_col=3)
    # labels
    labels = []
    r = labels_start_row
    while True:
        v = ws.cell(r, 2).value
        if stop_on_blank and _is_blank(v):
            break
        if _is_blank(v):
            labels.append("")
        else:
            labels.append(str(v).strip())
        r += 1
        # safety
        if r > ws.max_row + 50:
            break

    # build
    data = {}
    for i, d in enumerate(dates):
        col = 3 + i
        col_vals = []
        for rr in range(labels_start_row, labels_start_row + len(labels)):
            v = ws.cell(rr, col).value
            col_vals.append(0 if v is None else v)
        data[d] = col_vals

    df = pd.DataFrame(data, index=labels)
    df.index.name = None
    df = df.apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
    # drop empty label rows
    df = df[df.index.astype(str).str.strip() != ""]
    return df

def _ensure_labels(ws, labels_start_row: int, existing_labels: list[str], desired_labels: list[str]) -> dict[str, int]:
    """Assure les labels dans la colonne B.

    - Conserve l'ordre existant pour les labels d√©j√† pr√©sents.
    - Ajoute les labels manquants.
    - Si 'Total' existe dans le bloc, ins√®re les nouveaux labels JUSTE AVANT 'Total'
      (afin de garder 'Total' en dernier).
    """
    existing = [str(x).strip() for x in existing_labels if not _is_blank(x)]
    to_add = [lab for lab in desired_labels if lab not in existing]

    if to_add:
        # Position d'insertion : avant 'Total' si pr√©sent, sinon √† la fin du bloc
        total_pos = None
        for i, v in enumerate(existing):
            if v.lower() == "total":
                total_pos = i
                break

        if total_pos is None:
            insert_at = labels_start_row + len(existing)
            ws.insert_rows(insert_at, amount=len(to_add))
            for i, lab in enumerate(to_add):
                ws.cell(insert_at + i, 2).value = lab
            existing.extend(to_add)
        else:
            insert_at = labels_start_row + total_pos
            ws.insert_rows(insert_at, amount=len(to_add))
            for i, lab in enumerate(to_add):
                ws.cell(insert_at + i, 2).value = lab
            existing = existing[:total_pos] + to_add + existing[total_pos:]

    return {lab: labels_start_row + i for i, lab in enumerate(existing)}

def _write_series_to_block(ws, header_row: int, labels_start_row: int, snap: _date, values: pd.Series, desired_order: list[str] | None = None):
    # read existing labels
    existing_labels = []
    r = labels_start_row
    while True:
        v = ws.cell(r, 2).value
        if _is_blank(v):
            break
        existing_labels.append(str(v).strip())
        r += 1

    # decide desired labels
    if desired_order is None:
        desired_labels = list(existing_labels)
        # add any new labels at end
        for lab in list(values.index):
            lab_s = str(lab).strip()
            if lab_s and lab_s not in desired_labels:
                desired_labels.append(lab_s)
    else:
        desired_labels = desired_order.copy()
        for lab in list(values.index):
            lab_s = str(lab).strip()
            if lab_s and lab_s not in desired_labels:
                desired_labels.append(lab_s)

    mapping = _ensure_labels(ws, labels_start_row, existing_labels, desired_labels)
    col = _find_or_append_date_col(ws, header_row, snap, start_col=3)

    # write
    for lab, row in mapping.items():
        ws.cell(row, col).value = int(values.get(lab, 0))



def _pick_kpi_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    """Choisit la feuille qui contient les tableaux KPI.

    On √©vite d'utiliser wb.active car l'active peut changer apr√®s export/sauvegarde.
    Crit√®re : pr√©sence du titre du tableau 1 (et √† d√©faut tableau 3).
    """
    for name in wb.sheetnames:
        ws = wb[name]
        try:
            if _find_row_by_title(ws, _T1_TITLE) is not None:
                return ws
        except Exception:
            pass
        try:
            if _find_row_by_title(ws, _T3_TITLE) is not None:
                return ws
        except Exception:
            pass
    return wb.active

def _build_states_order(existing_labels: list[str], new_labels: list[str]) -> list[str]:
    """Construit un ordre d'√©tats bas√© sur l'historique (tableaux_hebdo) + nouveaux √©tats (CDC).
    R√®gle: tous les √©tats (existants + nouveaux) conservent l'ordre existant, les nouveaux sont ajout√©s
    juste AVANT 'Total', et 'Total' reste en dernier.
    """
    existing = [str(x).strip() for x in existing_labels if not _is_blank(x)]
    new = [str(x).strip() for x in new_labels if not _is_blank(x)]

    # D√©tecter Total (peut √™tre absent)
    has_total = any(x.lower() == "total" for x in existing + new)

    # Base = ordre existant sans Total
    base = [x for x in existing if x.lower() != "total"]

    # Ajouter les nouveaux labels (sans Total) qui n'existent pas encore
    for lab in new:
        if lab.lower() == "total":
            continue
        if lab not in base:
            base.append(lab)

    # Remettre Total en dernier si pr√©sent
    if has_total:
        base.append("Total")
    return base

def plot_stacked_bar(df: pd.DataFrame, title: str, drop_labels: set[str] | None = None):
    """Stacked bar chart aligned with the displayed KPI tables.

    - Handles empty tables gracefully
    - Normalizes/sorts date-like columns chronologically
    - Keeps non-date columns (e.g. a CDC snapshot) at the end
    """
    import plotly.graph_objects as go

    if df is None or df.empty:
        st.info(f"Aucune donn√©e √† tracer pour : {title}")
        return

    dff = df.copy()

    # Drop rows like 'Total' for charts when requested
    if drop_labels:
        dff = dff.drop(index=[x for x in drop_labels if x in dff.index], errors="ignore")

    if dff.empty:
        st.info(f"Aucune donn√©e √† tracer pour : {title}")
        return

    # Ensure numeric
    dff = dff.apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)

    # Normalize columns: parse date-like headers and sort chronologically
    col_info = []
    for c in list(dff.columns):
        dt = pd.to_datetime(c, errors="coerce")
        if pd.isna(dt):
            col_info.append((c, None))
        else:
            col_info.append((c, dt.date()))

    date_cols = [c for c, d in col_info if d is not None]
    other_cols = [c for c, d in col_info if d is None]

    # Sort only the date-like columns
    if date_cols:
        date_cols_sorted = sorted(date_cols, key=lambda c: pd.to_datetime(c, errors="coerce"))
        ordered_cols = date_cols_sorted + [c for c in other_cols if c not in date_cols_sorted]
        dff = dff.loc[:, ordered_cols]

    # X labels
    x = []
    for c in dff.columns:
        dt = pd.to_datetime(c, errors="coerce")
        if not pd.isna(dt):
            x.append(dt.strftime("%d/%m/%Y"))
        else:
            x.append(str(c))

    fig = go.Figure()
    for lab in dff.index:
        fig.add_trace(go.Bar(name=str(lab), x=x, y=dff.loc[lab].tolist()))
    fig.update_layout(barmode="stack", title=title, height=420, margin=dict(l=10, r=10, t=60, b=10))
    st.plotly_chart(fig, use_container_width=True)

# ========================= UI KPI =========================

# ‚úÖ √©viter NameError si l'historique n'est pas encore upload√©
tab1_hist = st.session_state.get("tab1_hist", pd.DataFrame())
tab2_hist = st.session_state.get("tab2_hist", pd.DataFrame())
tab3_hist = st.session_state.get("tab3_hist", pd.DataFrame())

st.title("KPI hebdomadaires")

with st.expander("üìå KPI (3 tableaux puis 3 graphiques)", expanded=True):
    cols = st.columns([1.2, 1.2, 1.2])
    cdc_file = cols[0].file_uploader("1) Uploader le cahier des charges (CDC)", type=["xlsx"], key="cdc_kpi")
    hist_file = cols[1].file_uploader("2) Uploader l'historique (tableaux_hebdo.xlsx)", type=["xlsx"], key="hist_kpi")
    snap = cols[2].date_input("3) Date de la semaine", value=_date.today(), key="snap_kpi")
    fiches_files = st.session_state.get("fiches_uploaded_files", None)  # fiches upload√©es en haut

        # ‚úÖ Persister l'historique pour √©viter qu'il "disparaisse" lors des reruns (upload fiches/CDC)
    if hist_file is not None:
        st.session_state["hist_bytes"] = hist_file.getvalue()

    hist_bytes = st.session_state.get("hist_bytes")

    if hist_bytes is None:
        pass
    else:
        hist_bytes = hist_file.getvalue()
        # Workbook pour LECTURE (valeurs) : √©vite de perdre les valeurs de formules apr√®s √©criture
        hist_wb_values = load_workbook(BytesIO(hist_bytes), data_only=True)
        ws_values = _pick_kpi_sheet(hist_wb_values)

        # Workbook pour √âCRITURE (formules conserv√©es) : utilis√© uniquement pour g√©n√©rer le fichier √† t√©l√©charger
        hist_wb_edit = load_workbook(BytesIO(hist_bytes), data_only=False)
        # utiliser la m√™me feuille que pour la lecture si possible
        try:
            ws_edit = hist_wb_edit[ws_values.title]
        except Exception:
            ws_edit = _pick_kpi_sheet(hist_wb_edit)

        # --- Parse CDC (si fourni)
        cdc_df = None
        if cdc_file is not None:
            try:
                cdc_df = parse_cdc_functions(cdc_file.getvalue())
            except Exception as e:
                st.error(str(e))

        t1_current = None
        t2_current = None


        # ===== Tableau 1 & 2 depuis CDC =====
        if cdc_df is not None and not cdc_df.empty:
            cdc_df["√âtat_norm"] = cdc_df["√âtat"].astype(str).str.strip()
            cdc_df["Cat√©gorie_norm"] = cdc_df["Cat√©gorie"].astype(str).str.strip()

            # Tableau 1: En test par BU (Cat√©gorie)
            en_test = cdc_df[cdc_df["√âtat_norm"].str.lower() == "en test"].copy()
            cats_order = ["IB","Commun sp√©cifique IB","Commun sp√©cifique PCM","Commun universel","PCM"]
            t1_current = en_test.groupby("Cat√©gorie_norm")["ID fonctionnalit√©"].nunique().reindex(cats_order).fillna(0).astype(int)

            # Tableau 2: √âtat des fonctionnalit√©s (comptage par √©tat)
            t2_current = (
                cdc_df.groupby("√âtat_norm")["ID fonctionnalit√©"]
                .nunique()
                .astype(int)
            )
            # Total = somme des √©tats (si pas d√©j√† pr√©sent)
            if "Total" not in t2_current.index:
                t2_current.loc["Total"] = int(t2_current.sum())

            # --- √âcrire dans l'historique (Tableau 1 & 2) sur le workbook d'√©dition
            t1_row = _find_row_by_title(ws_edit, _T1_TITLE) or 3
            _write_series_to_block(ws_edit, t1_row, t1_row + 1, snap, t1_current, desired_order=cats_order)

            # Ordre des √©tats = ordre existant dans tableaux_hebdo (B20..) + nouveaux √©tats du CDC,
            # en gardant 'Total' en dernier.
            t2_row, t2_labels_row = _normalize_table2_location(ws_values)
            tab2_existing = _read_block(ws_values, t2_row, t2_labels_row)
            dyn_states_order = _build_states_order(list(tab2_existing.index), list(t2_current.index))
            t2_row_edit, t2_labels_row_edit = _normalize_table2_location(ws_edit)
            _write_series_to_block(ws_edit, t2_row_edit, t2_labels_row_edit, snap, t2_current, desired_order=dyn_states_order)

        # ===== Tableau 3 depuis les FICHES (PCM / type de doc) =====
        # 4) Calculer t3_current uniquement √† partir des fiches valides
        # ===== Tableau 3 depuis les FICHES (PCM / type de doc) =====
        t3_current = pd.Series(dtype=int)

        dfs_kpi = []
        ignored = []

        if fiches_files:
            # 1) Prot√©ger l'it√©ration (zip corrompu, flux, etc.)
            try:
                excel_files = list(iter_excel_files(fiches_files))
            except Exception as e:
                st.error(f"Erreur en lisant les fiches (zip/lecture): {e}")
                excel_files = []

            # 2) Prot√©ger le parsing fiche par fiche
            for f in excel_files:
                name = str(getattr(f, "name", ""))

                # ignore r√©f√©rentiel si pr√©sent dans l'upload
                if name.lower().startswith("pourscript-tableauxjeremie"):
                    continue

                try:
                    df = parse_fiche(f)
                    if df is not None and not df.empty:
                        dfs_kpi.append(df)
                except Exception as e:
                    ignored.append((name or "fichier_sans_nom", str(e)))

            # 3) Afficher les ignor√©s (optionnel)
            if ignored:
                with st.expander(f"‚ö†Ô∏è Fiches ignor√©es ({len(ignored)})", expanded=False):
                    for n, err in ignored[:50]:
                        st.write(f"- {n}: {err}")
                    if len(ignored) > 50:
                        st.write(f"... +{len(ignored)-50} autres")

            # 4) Calculer t3_current uniquement √† partir des fiches valides
            if dfs_kpi:
                data_kpi = pd.concat(dfs_kpi, ignore_index=True)
                fiches = data_kpi.drop_duplicates(subset="fiche_name").copy()

                fiches_pcm = fiches[fiches["site"].isin(["STMA", "JMT", "FLC"])].copy()
                fiches_pcm["type_document"] = fiches_pcm["type_document"].astype(str).str.strip()
                fiches_pcm.loc[fiches_pcm["type_document"].isin(["", "None", "nan"]), "type_document"] = "Inconnu"

                t3_current = (
                    fiches_pcm.groupby("type_document")["fiche_name"]
                    .nunique()
                    .sort_index()
                    .astype(int)
                )

        # --- √âcrire dans l'historique (Tableau 3)
        if t3_current is not None and not t3_current.empty:
            t3_row = _find_row_by_title(ws_edit, _T3_TITLE) or 12
            _write_series_to_block(ws_edit, t3_row, t3_row+1, snap, t3_current, desired_order=None)



        # ===== Lire les 3 tableaux depuis l'historique (apr√®s √©criture) =====
        # Tableau 1
        t1_row = _find_row_by_title(ws_values, _T1_TITLE) or 3
        tab1_hist = _read_block(ws_values, t1_row, t1_row+1)

        # Tableau 2
        t2_row, t2_labels_row = _normalize_table2_location(ws_values)
        tab2_hist = _read_block(ws_values, t2_row, t2_labels_row)
        # pour le tableau 2 on veut garder Total dans le tableau
        # (mais on le supprimera sur le graphique)
        # Assure l'ordre bas√© sur l'historique (et garde Total en dernier si pr√©sent)
        dyn_states_order = _build_states_order(list(tab2_hist.index), list(tab2_hist.index))
        tab2_hist = tab2_hist.reindex([s for s in dyn_states_order if s in tab2_hist.index]).fillna(0)
        tab2_hist = tab2_hist.apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)

        # Tableau 3
        t3_row = _find_row_by_title(ws_edit, _T3_TITLE) or 12
        tab3_hist = _read_block(ws_values, t3_row, t3_row+1)

        
        # ===== Construire les tableaux √† AFFICHER : base = tableaux_hebdo (valeurs), + derni√®re colonne = CDC si fourni
        tab1_display = tab1_hist.copy()
        tab2_display = tab2_hist.copy()
        tab3_display = tab3_hist.copy()

        if t1_current is not None and hasattr(t1_current, "empty") and (not t1_current.empty):
            col_name = snap.strftime("%Y-%m-%d")
            tab1_display[col_name] = t1_current.reindex(tab1_display.index).fillna(0).astype(int)

        if t2_current is not None and hasattr(t2_current, "empty") and (not t2_current.empty):
            col_name = snap.strftime("%Y-%m-%d")
            # √âtendre l'index avec les nouveaux √©tats du CDC, en gardant 'Total' en dernier
            dyn_states_order = _build_states_order(list(tab2_display.index), list(t2_current.index))
            tab2_display = tab2_display.reindex(dyn_states_order).fillna(0).astype(int)
            # Ajouter la colonne CDC align√©e sur le nouvel index (nouveaux √©tats inclus)
            tab2_display[col_name] = t2_current.reindex(tab2_display.index).fillna(0).astype(int)

        # Tableau 3 : derni√®re colonne = FICHES (si fournies). C'est le seul tableau d√©pendant des fiches.
        if t3_current is not None and hasattr(t3_current, "empty") and (not t3_current.empty):
            col_name = snap.strftime("%Y-%m-%d")
            # Conserver exactement la base tableaux_hebdo, mais ajouter/mettre √† jour la derni√®re colonne.
            # Si de nouveaux types apparaissent, on les ajoute en bas.
            new_index = tab3_display.index.union(t3_current.index)
            tab3_display = tab3_display.reindex(new_index).fillna(0)
            tab3_display[col_name] = t3_current.reindex(tab3_display.index).fillna(0).astype(int)

        # Utiliser les versions enrichies pour l'affichage/graphiques
        tab1_hist = tab1_display
        tab2_hist = tab2_display
        tab3_hist = tab3_display

# ‚úÖ Persistance : conserver les tableaux m√™me lors des reruns Streamlit
        st.session_state["tab1_hist"] = tab1_display
        st.session_state["tab2_hist"] = tab2_display
        st.session_state["tab3_hist"] = tab3_display

        # download updated history
        out = BytesIO()
        hist_wb_edit.save(out)
        st.download_button(
            "T√©l√©charger l'historique mis √† jour (tableaux_hebdo.xlsx)",
            data=out.getvalue(),
            file_name=f"tableaux_hebdo_{snap:%d-%m-%y}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_hist_kpi"
        )
    # ------------------------------------------------------------
    # Tableaux & graphiques KPI (masquables)
    # ------------------------------------------------------------
    if (hist_file is not None) or (not tab1_hist.empty) or (not tab2_hist.empty) or (not tab3_hist.empty):
        st.markdown("### Tableaux")

        st.markdown("**1) Nombre de fonctionnalit√©s en phase de test par BU**")
        st.dataframe(tab1_hist, use_container_width=True)

        st.markdown("**2) √âtat des fonctionnalit√©s pr√©sentes dans le Cahier des Charges**")
        st.dataframe(tab2_hist, use_container_width=True)

        st.markdown("**3) Nombre de Fiches de Test remplies PCM par type de document**")
        st.dataframe(tab3_hist, use_container_width=True)

        # ------------------------------------------------------------
        # üìå Graphiques KPI ‚Äî Toujours visibles et ind√©pendants des fiches
        # ------------------------------------------------------------
        st.markdown("### Graphiques")

        plot_stacked_bar(
            tab1_hist,
            "1) Nombre de fonctionnalit√©s en phase de test par BU"
        )

        plot_stacked_bar(
            tab2_hist,
            "2) √âtat des fonctionnalit√©s pr√©sentes dans le Cahier des Charges",
            drop_labels={"Total"}
        )

        plot_stacked_bar(
            tab3_hist,
            "3) Nombre de Fiches de Test remplies PCM par type de document"
        )


    else:
        st.info("Charge `tableaux_hebdo.xlsx` pour afficher les tableaux et graphiques KPI.")
st.markdown("---")

#Nettoie les cellules des fiches de test
def _normalize_text(s: object) -> str:
    """Normalise un texte (minuscule, sans accents, espaces simplifi√©s)."""
    if s is None:
        return ""
    s= str(s).strip().lower()
    # Retire les accents
    s=unicodedata.normalize("NFKD", s)
    s= "".join(ch for ch in s if not unicodedata.combining(ch))
    s= re.sub(r"\s+", " ", s)
    return s

def verdict_to_score(verdict: object) -> float:
    """Mappe strictement le verdict vers un score : bon=1, partiel=0.5, mauvais=0."""
    if not isinstance(verdict, str):
        return np.nan
    
    v = verdict.strip() # On garde juste le retrait des espaces inutiles
    
    if v == "Bon":
        return 1.0
    if v == "Partiellement bon":
        return 0.5
    if v == "Mauvais":
        return 0.0
        
    return np.nan # Si c'est vide ou √©crit autrement

# ---------------------------------------------------------------------
def infer_fo_cols(df: pd.DataFrame) -> list[str]:
    """D√©tecte automatiquement les colonnes fonctions outils (fo1, fo2, ...), tri√©es num√©riquement"""
    cols = []
    for c in df.columns:
        cs = str(c).strip()
        if re.match(r"^fo\d+$", cs):
            cols.append(cs)
    def _k(x: str) -> int:
        m = re.search(r"(\d+)$", x)
        return int(m.group(1)) if m else 0
    return sorted(cols, key=_k)


def excel_sheet_name(title: str) -> str:
    """Nom des tableaux-> feuilles pour export excel (m√™me noms que ceux des tableaux affich√©s sur le streamlit) : 
    - max 31 caract√®res
    - suppression des caract√®res interdits : : \ / ? * [ ]
    """
    if not isinstance(title, str):
        title = str(title)
    # Caract√®res interdits Excel
    title = re.sub(r"[:\\/\?\*\[\]]", "", title)
    title = title.strip().strip("'")
    if not title:
        title = "Feuil1"
    return title[:31]


# =====================================================================
# STYLES / COULEURS
# =====================================================================

#Couleurs pour les verdicts
def color_verdict(val):
    if not isinstance(val, str):
        return ""
    v = val.strip().lower()
    if v == "bon":
        return "background-color: #c6efce; color: #006100; font-weight: bold;"
    if "partiellement" in v:
        return "background-color: #ffe699; color: #7f6000; font-weight: bold;"
    if "mauvais" in v:
        return "background-color: #fd4a4a; color: #9c0006; font-weight: bold;"
    return ""

# Regarde dans le tableau avec le nb de tests et performance √† atteindre
def extract_pct(cell):
    if not isinstance(cell, str):
        return None
    cell = cell.strip()
    if not cell or cell in ("tbd", "NA"):
        return None
    if "(" not in cell or "%" not in cell:
        return None
    try:
        inside = cell.split("(")[1].split(")")[0]   # ex: '0.7%' ou '25.0%'
        inside = inside.replace("%", "").replace(",", ".")
        return float(inside)
    except Exception:
        return None


def color_tab3(val):
    pct = extract_pct(val)
    if pct is None:
        return ""
    # pct est en pourcentage (0‚Äì100)
    if pct < 25:
        return "background-color: #fd4a4a;"               # rouge
    elif pct < 50:
        return "background-color: #f4b183;"               # orange
    elif pct < 75:
        return "background-color: #fff2cc;"               # jaune
    elif pct < 90:
        return "background-color: #c6efce;"               # vert clair
    else:
        return "background-color: #00b050; color: white;" # vert fonc√©


#tableau justesse
def build_tab2bis(data, ref_base):
    #seulement les fs r√©ellement test√©s
    fs_testes = data["fs_id"].astype(str).str.strip().unique()
    tab2bis = ref_base[ref_base["N¬∞ fs"].isin(fs_testes)].copy()

    #Calcul de la moyenne des scores (0 √† 100) avec bon=100%, partiel=50%, mauvais=0%
    tmp = data.copy()
    tmp["__score__"] = tmp["verdict_doc"].apply(verdict_to_score) * 100.0
    taux_fs = (
        tmp.groupby("fs_id")["__score__"]
        .mean()
        .rename(lambda x: str(x).strip())
        .to_dict()
    )

    # Remplit les cases avec le % calcul√©
    fo_cols = infer_fo_cols(ref_base)
    def fill_cell(row, fo):
        val = row[fo]
        if val != "x":
            return ""
        fs = row["N¬∞ fs"]
        actual = taux_fs.get(fs)
        if actual is None or pd.isna(actual):
            return ""
        return f"{actual:.1f}%"

    for fo in fo_cols:
        tab2bis[fo] = tab2bis.apply(lambda r: fill_cell(r, fo), axis=1)

    return tab2bis, fo_cols


def style_tab2bis(df, fo_cols):
    """Applique les couleurs sur le Tableau Justesse."""
    styles = _DataFrame("", index=df.index, columns=df.columns)

    for i in df.index:
        for fo in fo_cols:
            val = df.at[i, fo]
            if not isinstance(val, str) or not val:
                continue
            if val in ("tbd", "NA"):
                continue
            try:
                pct = float(val.replace("%", "").replace(",", "."))
            except Exception:
                continue
            #en fonction du pourcentage    
            if pct < 25:
                styles.at[i, fo] = "background-color: #fd4a4a;"
            elif pct < 50:
                styles.at[i, fo] = "background-color: #f4b183;"
            elif pct < 75:
                styles.at[i, fo] = "background-color: #fff2cc;"
            elif pct < 90:
                styles.at[i, fo] = "background-color: #c6efce;"
            else:
                styles.at[i, fo] = "background-color: #00b050; color: white;"

    return styles


def norm_verdict(x: str) -> str:
    """Simplifie le verdict pour les statistiques (compter combien de 'bon', 'mauvais'...)"""
    if not isinstance(x, str):
        return "none"
    x = x.strip().lower()
    if x in ("", "nan", "none"):
        return "none"
    if x == "bon":
        return "bon"
    if "partiel" in x:
        return "partiel"
    if "mauvais" in x:
        return "mauvais"
    return "none"



def extract_modification(ref):
    """Extrait le code modification (ex: m01) depuis la r√©f√©rence."""
    if not isinstance(ref, str):
        return ""
    ref = ref.strip()
    match = re.search(r"(m\d{2})$", ref, re.IGNORECASE)
    return match.group(1) if match else ""


def extract_doc_id(ref: str) -> str:
    """Extrait un identifiant de document depuis la r√©f√©rence (nom de fichier).
    Heuristique : cherche d'abord un motif explicite (doc/document/coedm/ref + chiffres),
    sinon prend la premi√®re s√©quence de >=4 chiffres, sinon retourne la r√©f√©rence enti√®re.
    """
    if not isinstance(ref, str):
        return ""
    s = ref.strip()
    m = re.search(r"(?:doc|document|coedm|ref)[-_ ]*(\d+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4,})", s)
    if m:
        return m.group(1)
    return s



# =====================================================================
# PARSE FICHE
# =====================================================================

#permet de lire les cases des fiches de test
def parse_fiche(file):
    filename = file.name

    if filename.startswith("~$"):
        return pd.DataFrame()  # fichier temporaire Excel => on ignore (j'ai d√©ja rencontr√© le cas, d'o√π cette s√©curit√©)

    # Exemples de noms :
    # - fs60-IA-v00-TUF-FFP_PV_...-Fiche-v01-CM.xlsx
    # - fs1-IA-v00-NR-...-Fiche-v01-CM.xlsx (fs 1 √† 3 chiffres)
    filename_clean = str(filename).strip()

    # --- 1. Analyse du NOM du fichier ---
    # V√©rifie si le nom respecte le format strict fsXX-IA-vXX..
    pat = r"^(fs\d{1,3})-IA-v\d+-(TUF|NR)-(.+)-Fiche-v\d+-[A-Za-z]+\.xlsx$" # si le nom des fiches change, c'est ici qu'il faut faire les modifs !
    m = re.match(pat, filename_clean, flags=re.IGNORECASE)
    if not m:
        # Si le nom est mauvais, le script s'arr√™te ici pour ce fichier
        raise ValueError(
            "Nom de fichier invalide. Format attendu : "
            "fs<1-3 chiffres>-IA-vXX-(TUF|NR)-<ref>-Fiche-vXX-<initiales>.xlsx"
        )

    fs_id = m.group(1).lower()
    fiche_type = m.group(2).upper()
    ref_from_filename = m.group(3)
    doc_id = extract_doc_id(ref_from_filename)

    # apr√®s upload des fiches de test, l'outil va lire la feuille "Template Fiche de Test" <- si le nom de la feuille excel change c'est ici qu'il faut modifier 
    df = pd.read_excel(file, sheet_name="Template Fiche de Test", header=None)

    # -----------------------------------------------------------------
    # D√©tection du d√©calage : ligne 18 contient-elle "*Searchable:" ? (apr√®s la nouvelle version du template des fiches de test)
    # -----------------------------------------------------------------
    has_searchable_row = False
    try:
        row18 = df.iloc[17]  # ligne 18 (index 17)
        for val in row18:
            if isinstance(val, str) and "Searchable" in val:
                has_searchable_row = True
                break
    except Exception:
        pass

    offset = 0 if has_searchable_row else -1

    def r(base_row):
        # applique le d√©calage seulement apr√®s la ligne 18
        if base_row > 17:
            return base_row + offset
        return base_row

    # ===========================
    # Champs "avant 18" R√©cup√©ration des infos g√©n√©rales (En-t√™te de fiche)
    # ===========================
    type_test = df.iloc[4, 1] if pd.notna(df.iloc[4, 1]) else None
    type_doc = df.iloc[14, 1] if pd.notna(df.iloc[14, 1]) else None
    ref_coedm = ref_from_filename

    # Label fonctionnalit√© (ligne 24 ‚Üí index 23, apr√®s 18 ‚Üí offset)
    label_fonctionnalite = None
    try:
        val_lbl = df.iloc[r(23), 1]
        if pd.notna(val_lbl):
            label_fonctionnalite = val_lbl
    except Exception:
        pass

    if label_fonctionnalite is None or (
        isinstance(label_fonctionnalite, float) and pd.isna(label_fonctionnalite)
    ):
        label_fonctionnalite = fs_id

    # Date (B10 ‚Üí ligne 10, index 9) = avant 18
    date_test = None
    try:
        val_date = df.iloc[9, 1]
        if not pd.isna(val_date):
            date_test = pd.to_datetime(val_date, errors="coerce")
    except Exception:
        pass

    # Site (B9 ‚Üí ligne 9, index 8, colonne 1)
    site = None
    try:
        val_site = df.iloc[8, 1]  # B9
        if not pd.isna(val_site):
            site = str(val_site).strip()
            if site == "LTA":
                site = "STMA"
    except Exception:
        pass

    # Code modification (B22)
    code_modif = None
    try:
        val_modif = df.iloc[21, 1]  # B22
        if isinstance(val_modif, str) and val_modif.strip():
            code_modif = val_modif.strip()
    except Exception:
        pass
    

    # Classe documentaire (D15 ‚Üí ligne 15, index 14) = avant 18
    classe_documentaire = None
    try:
        val_cls = df.iloc[14, 3]
        if pd.isna(val_cls):
            classe_documentaire = None
        elif isinstance(val_cls, str):
            v = val_cls.strip()
            if not v:
                classe_documentaire = None
            elif v.lower() == "non":
                classe_documentaire = "Non Searchable"
            elif v.lower() == "oui":
                classe_documentaire = "Searchable"
            else:
                classe_documentaire = v
        else:
            classe_documentaire = val_cls
    except Exception:
        pass

    # Nombre de pages total (ligne 17 ‚Üí index 16)
    try:
        nb_pages_total = int(df.iloc[16, 1])
    except Exception:
        nb_pages_total = None

    # Commentaire additionnel (ligne 46 ‚Üí index 45, apr√®s 18 ‚Üí offset)
    commentaire_add = None
    try:
        comment_row = df.iloc[r(45)]
        for val in comment_row:
            if isinstance(val, str) and val.strip():
                commentaire_add = val.strip()
                break
    except Exception:
        pass

    # Verdict doc : lignes 6 √† 12 (index 5..11)
    verdict_doc = None
    try:
        for rr in range(5, 12):
            val = df.iloc[rr, 2]
            if isinstance(val, str) and val.strip():
                verdict_doc = val.strip()
                break
    except Exception:
        pass

    verdict_score = verdict_to_score(verdict_doc)

    # Nom testeur : lignes 6 et 7 (index 5 & 6)
    nom_testeur = None
    try:
        val7 = df.iloc[6, 1]
        val6 = df.iloc[5, 1]
        out = []
        if isinstance(val7, str) and val7.strip():
            out.append(val7.strip())
        if isinstance(val6, str) and val6.strip():
            out.append(val6.strip())
        if out:
            nom_testeur = " ".join(out)
    except Exception:
        pass

    # Fonctionnalit√© : 25A / 26A avec offset
    fonctionnalite = None
    try:
        cell_26A = df.iloc[r(25), 0]  # ligne 26 (index 25)
        cell_25A = df.iloc[r(24), 0]  # ligne 25 (index 24)
        sentinel = "[NR] Nombre de test de r√©p√©tabilit√© requis"

        if isinstance(cell_26A, str) and cell_26A.strip() and cell_26A.strip() != sentinel:
            fonctionnalite = cell_26A.strip()
        elif isinstance(cell_25A, str) and cell_25A.strip():
            fonctionnalite = cell_25A.strip()
    except Exception:
        pass

    # Noms des tests (ligne 31 ‚Üí index 30, apr√®s 18 ‚Üí offset)
    try:
        tests = list(df.iloc[r(30), 2:6].dropna())
    except Exception:
        tests = []

    records = []
    # Pour chaque test trouv√©...
    for k in range(len(tests)):
        col = 2 + k # D√©calage de colonnes (C=2, D=3...)

        # lignes 32..37 ‚Üí index 31..36 ‚Üí apr√®s 18 ‚Üí offset
        # On r√©cup√®re les lignes de r√©sultats (Temps, FN, FP...)
        row_h = r(31)
        row_m = r(32)
        row_j = r(33)
        row_fn = r(34)
        row_fp = r(35)
        row_inc = r(36)

        raw_vals = [
            df.iloc[row_h, col],   # humain
            df.iloc[row_m, col],   # machine
            df.iloc[row_j, col],   # justes
            df.iloc[row_fn, col],  # fn
            df.iloc[row_fp, col],  # fp
            df.iloc[row_inc, col], # incertaines
        ]

        # Test totalement vide ‚Üí ignor√©
        if all(pd.isna(v) for v in raw_vals):
            continue

        def to_float_or_none(v):
            try:
                return float(v) if not pd.isna(v) else None
            except Exception:
                return None

        humain = to_float_or_none(raw_vals[0])
        machine = to_float_or_none(raw_vals[1])
        justes = to_float_or_none(raw_vals[2])
        fn = to_float_or_none(raw_vals[3])
        fp = to_float_or_none(raw_vals[4])
        inc = to_float_or_none(raw_vals[5])

        # Cr√©ation de l'objet r√©sultat
        rec = {
            "fs_id": fs_id,
            "ref_coedm": ref_coedm,
            "doc_id": doc_id,
            "fiche_type": fiche_type,
            "fiche_name": filename,
            "date_test": date_test,
            "type_test": type_test,
            "type_document": type_doc,
            "classe_documentaire": classe_documentaire,
            "site": site,
            "label_fonctionnalite": label_fonctionnalite,
            "code_modif": code_modif,
            "nb_pages_total": nb_pages_total,
            "commentaire_additionnel": commentaire_add,
            "verdict_doc": verdict_doc,
            "verdict_score": verdict_score,
            "nom_testeur": nom_testeur,
            "fonctionnalite": fonctionnalite,
            "test_label": f"Test {k+1}",
            "temps_humain_s": humain,
            "temps_machine_s": machine,
            "nb_pages_justes": justes,
            "fn": fn,
            "fp": fp,
            "incertaines": inc,
        }
        # Calcul des pourcentages
        if nb_pages_total and nb_pages_total != 0:
            rec["taux_justes"] = (justes / nb_pages_total * 100) if justes is not None else None
            rec["taux_fn"] = (fn / nb_pages_total * 100) if fn is not None else None
            rec["taux_fp"] = (fp / nb_pages_total * 100) if fp is not None else None
            rec["taux_incertaines"] = (inc / nb_pages_total * 100) if inc is not None else None
        else:
            rec["taux_justes"] = None
            rec["taux_fn"] = None
            rec["taux_fp"] = None
            rec["taux_incertaines"] = None


        if humain is not None and machine is not None:
            rec["gain_temps_s"] = humain - machine
        else:
            rec["gain_temps_s"] = None

        records.append(rec)

    return pd.DataFrame(records)


# =====================================================================
# INTERFACE STREAMLIT
# =====================================================================

st.title("Analyse automatique des fiches IA4Doc")

# Zone de glisser-d√©poser des fichiers
uploaded_files = st.session_state.get("fiches_uploaded_files", None)  # utilise l‚Äôuploader du haut


if uploaded_files:

    # --- Chargement et Analyse des donn√©es ---
    # On stocke les erreurs dans la session pour pouvoir les supprimer (bouton ‚ùå)
    if "file_errors" not in st.session_state:
        st.session_state["file_errors"] = {}

    # Fichiers √† ignorer (pour masquer une erreur sans re-uploader)
    if "ignored_fiche_filenames" not in st.session_state:
        st.session_state["ignored_fiche_filenames"] = set()

    @st.cache_data(show_spinner=False)
    def _parse_fiche_cached(fname: str, file_bytes: bytes):
        """Parse une fiche en cache pour √©viter de tout retraiter √† chaque rerun."""
        bio = BytesIO(file_bytes)
        # pd.read_excel accepte un file-like; on garde le nom pour la validation du pattern
        try:
            bio.name = fname  # type: ignore[attr-defined]
        except Exception:
            pass
        return parse_fiche(bio)


    def add_file_error(fname: str, msg: str):
        # Garde la derni√®re erreur pour ce fichier (simple)
        st.session_state["file_errors"][fname] = msg

    # ------------------------------------------------------
    # PARSE DES FICHES + anti-doublon
    # ------------------------------------------------------
    dfs = []
    seen_files = set()

    try:
        excel_files = list(iter_excel_files(uploaded_files))
    except Exception as e:
        st.error(f"Erreur ZIP/lecture: {e}")
        excel_files = []

    for file in excel_files:

        # Anti-doublon de fichier
        if file.name in seen_files:
            add_file_error(file.name, f"Fiche d√©j√† charg√©e : {file.name} (ignor√©e)")
            continue
        seen_files.add(file.name)

        # Le fichier r√©f√©rentiel (pourScript-tableauxJeremie.xlsx) n'est pas une fiche de test
        if file.name.lower().startswith("pourscript-tableauxjeremie"):
            continue

        # Si l'utilisateur a "supprim√©" l'erreur (‚ùå), on ignore compl√®tement ce fichier
        if file.name in st.session_state["ignored_fiche_filenames"]:
            continue


        try:
            # APPEL AU LECTEUR (parse_fiche) - version cache pour acc√©l√©rer les reruns
            try:
                file.seek(0)
            except Exception:
                pass
            try:
                file_bytes = file.getvalue()  # UploadedFile
            except Exception:
                file_bytes = file.read()
            df = _parse_fiche_cached(file.name, file_bytes)
            if not df.empty:
                dfs.append(df)
        except Exception as e:
            add_file_error(file.name, str(e))

    # Affiche les erreurs et permet de les enlever (juste l'affichage, pas besoin de re-uploader)
    if st.session_state["file_errors"]:
        st.subheader("Erreurs d√©tect√©es")
        for fname, msg in list(st.session_state["file_errors"].items()):
            c1, c2 = st.columns([12, 1])
            c1.error(f"{fname} : {msg}")
            if c2.button("‚ùå", key=f"rm_err_{fname}"):
                st.session_state["ignored_fiche_filenames"].add(fname)
                del st.session_state["file_errors"][fname]
                st.rerun()

    if not dfs:
        st.warning("Aucune fiche valide (les fiches en erreur ont √©t√© ignor√©es).")
        st.session_state["data_fiches"] = None
        # IMPORTANT: ne pas stopper l'app -> sinon Tableau/Graph 2 disparaissent
    else:
        data = pd.concat(dfs, ignore_index=True)

        # ID simple pour pouvoir exclure des lignes plus tard (1 ligne = 1 test)
        data = data.reset_index(drop=True)
        data["test_uid"] = data.index.astype(int)

        # ‚úÖ stocker APR√àS ajout de test_uid
        st.session_state["data_fiches"] = data

    
    data = st.session_state.get("data_fiches")

    # Si aucune fiche valide, on affiche juste les erreurs et on n'ex√©cute pas la suite,
    # mais on ne plante pas => les KPI (Tableau/Graph 2) restent affich√©s.
    if data is None or data.empty:
        st.info("Analyse fiches : aucune fiche valide √† traiter (les autres ont √©t√© ignor√©es).")
    else:
        # ------------------------------------------------------
        # NR / TUF : r√®gle m√©tier
        # ------------------------------------------------------
        # - Pour la PERF : si NR existe -> NR gagne ; si plusieurs NR -> on garde la plus r√©cente
        # - Si pas de NR : on garde la plus r√©cente des TUF
        # - IMPORTANT : on ne jette rien c√¥t√© volume. Tous les tests restent dans "data".
        #
        # Donc ici on fabrique juste un sous-jeu "data_calc" utilis√© pour les moyennes de score.
        overrides_report = pd.DataFrame()
        data["date_test"] = pd.to_datetime(data.get("date_test"), errors="coerce")

    def _pick_latest_fiche(g: pd.DataFrame, fiche_type: str):
        # Renvoie le nom de fiche la plus r√©cente pour un type donn√© (NR ou TUF)
        sub = g[g["fiche_type"].astype(str).str.upper() == fiche_type].copy()
        if sub.empty:
            return None
        # Une fiche = plusieurs lignes (Test 1, Test 2...), donc on d√©duplique au niveau fichier
        sub = sub.sort_values("date_test", ascending=False)
        return sub["fiche_name"].iloc[0]

    def _pick_best_fiche(g: pd.DataFrame) -> str:
        # D'abord NR (le plus r√©cent), sinon TUF (le plus r√©cent), sinon le plus r√©cent tout court
        fnr = _pick_latest_fiche(g, "NR")
        if fnr:
            return fnr
        ftuf = _pick_latest_fiche(g, "TUF")
        if ftuf:
            return ftuf
        g = g.sort_values("date_test", ascending=False)
        return g["fiche_name"].iloc[0]

    rows = []
    best_fiches = []

    if all(c in data.columns for c in ["doc_id", "fs_id", "fiche_type", "fiche_name"]):
        for (doc, fs), g in data.groupby(["doc_id", "fs_id"], dropna=False):
            best = _pick_best_fiche(g)
            best_fiches.append(best)

            # Petit rapport NR vs TUF (utile en audit)
            fnr = _pick_latest_fiche(g, "NR")
            ftuf = _pick_latest_fiche(g, "TUF")
            if fnr and ftuf:
                # score = verdict_score (d√©j√† entre 0 et 1)
                nr_score = pd.to_numeric(g[g["fiche_name"] == fnr]["verdict_score"], errors="coerce").mean()
                tuf_score = pd.to_numeric(g[g["fiche_name"] == ftuf]["verdict_score"], errors="coerce").mean()

                nr_verdict = g[g["fiche_name"] == fnr]["verdict_doc"].dropna().astype(str)
                tuf_verdict = g[g["fiche_name"] == ftuf]["verdict_doc"].dropna().astype(str)

                rows.append({
                    "doc_id": doc,
                    "fs_id": fs,
                    "fiche_NR": fnr,
                    "fiche_TUF": ftuf,
                    "verdict_NR": nr_verdict.mode().iloc[0] if len(nr_verdict) else "",
                    "verdict_TUF": tuf_verdict.mode().iloc[0] if len(tuf_verdict) else "",
                    "perf_NR": (nr_score * 100) if pd.notna(nr_score) else np.nan,
                    "perf_TUF": (tuf_score * 100) if pd.notna(tuf_score) else np.nan,
                    "delta_NR_minus_TUF": ((nr_score - tuf_score) * 100) if pd.notna(nr_score) and pd.notna(tuf_score) else np.nan,
                })

        if rows:
            overrides_report = pd.DataFrame(rows)

        # data_calc = uniquement les fiches "retenues" pour les calculs de perf
        data_calc = data[data["fiche_name"].isin(best_fiches)].copy()
    else:
        data_calc = data.copy()

    data["fs_id"] = data["fs_id"].astype(str).str.strip()


    # ------------------------------------------------------
    # Gestion des exclusions (tests √† ignorer)
    # ------------------------------------------------------

    # --- Gestion des exclusions manuelles (Cocher pour ignorer) ---
    if "excluded_test_uids" not in st.session_state:
        st.session_state["excluded_test_uids"] = set()

    # ‚úÖ S√©curit√© : test_uid doit exister quoi qu'il arrive
    if "test_uid" not in data.columns:
        data = data.reset_index(drop=True)
        data["test_uid"] = data.index.astype(int)
        st.session_state["data_fiches"] = data
    excluded = st.session_state["excluded_test_uids"]
    data_f = data[~data["test_uid"].isin(excluded)].copy()

    # M√™me logique de filtre mais pour les calculs de perf (NR plus r√©cent, etc.)
    # On applique aussi les exclusions manuelles.
    data_calc_f = data_calc[~data_calc["test_uid"].isin(excluded)].copy()

    # Colonnes "donn√©es extraites" (utilis√©es aussi pour les exports)
    clean_cols_internal = [
        "test_uid",
        "fs_id",
        "doc_id",
        "fiche_type",
        "fiche_name",
        "ref_coedm",
        "date_test",
        "nom_testeur",
        "type_test",
        "type_document",
        "classe_documentaire",
        "site",
        "fonctionnalite",
        "code_modif",
        "nb_pages_total",
        "test_label",
        "temps_humain_s",
        "temps_machine_s",
        "nb_pages_justes",
        "fn",
        "fp",
        "incertaines",
        "taux_justes",
        "taux_fn",
        "taux_fp",
        "taux_incertaines",
        "gain_temps_s",
        "verdict_doc",
    ]

    clean_cols = [c for c in clean_cols_internal if c not in ("test_uid", "doc_id")]

    # ==================================================================
    # CHARGEMENT R√âF√âRENTIEL (Feuil1 / Feuil2)
    # ==================================================================

    # C'est l√† qu'on utilise "infer_fo_cols" pour trouver les colonnes fo1...
    ref_xls = pd.ExcelFile("pourScript-tableauxJeremie.xlsx") # tableau excel √† fournir pour le r√©f√©rentiel des fonctionnalit√©s outil
    ref1 = ref_xls.parse("Feuil1")
    ref2 = ref_xls.parse("Feuil2")

    # Nettoyage colonnes vides
    ref1 = ref1.loc[:, ~ref1.columns.str.contains("Unnamed")]
    ref2 = ref2.loc[:, ~ref2.columns.str.contains("Unnamed")]

    # Unifier la colonne complexit√© (compatibilit√© anciennes versions)
    comp1 = ref2["complexit√©.1"] if "complexit√©.1" in ref2.columns else None
    comp0 = ref2["complexit√©"] if "complexit√©" in ref2.columns else None

    if comp1 is not None and comp0 is not None:
        ref2["complexit√©_unifiee"] = comp1.where(comp1.notna(), comp0)
    elif comp1 is not None:
        ref2["complexit√©_unifiee"] = comp1
    else:
        ref2["complexit√©_unifiee"] = comp0

    ref2["complexit√©_unifiee"] = ref2["complexit√©_unifiee"].fillna("tbd")
    ref2 = ref2.drop(columns=[c for c in ["complexit√©", "complexit√©.1"] if c in ref2.columns])
    ref2 = ref2.rename(columns={"complexit√©_unifiee": "complexit√©"})

    # Merge r√©f√©rentiel
    ref_full = ref1.merge(ref2, on="N¬∞ fs", how="left")
    ref_full["N¬∞ fs"] = ref_full["N¬∞ fs"].astype(str).str.strip()

    # Colonnes FO d√©tect√©es automatiquement (fo1..foN)
    fo_cols = infer_fo_cols(ref_full)

    # On ajoute fo0 si elle n'existe pas
    if "fo0" not in fo_cols:
        ref_full["fo0"] = ""
        fo_cols = ["fo0"] + fo_cols


    base_cols = ["N¬∞ fs"] + fo_cols + ["complexit√©"]
    ref_base = ref_full[base_cols].copy()

    # Assure qu'il y a au moins une colonne coch√©e 'x' 
    def ensure_at_least_one_fo(row):
    # Si aucune FO coch√©e, on met fo0
        if not any(str(row.get(fo, "")).strip().lower() == "x" for fo in fo_cols):
            row["fo0"] = "x"
        return row


    ref_base = ref_base.apply(ensure_at_least_one_fo, axis=1)

    # ==================================================================
    # Crit√®res / nb tests : depuis Excel si disponible, sinon fallback
    # ==================================================================


    #A revoir
    # --- R√©cup√©ration des Crit√®res de R√©ussite (99% etc) ---
    # Fonctions helpers pour lire les %, int ou 'tbd'
    def _to_percent_str(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "tbd"
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return "tbd"
            # d√©j√† "99%" ?
            if "%" in s:
                return s
            # "0.99" ?
            try:
                f = float(s.replace(",", "."))
                if 0 <= f <= 1.2:
                    return f"{f*100:.0f}%"
                if 1.2 < f <= 100:
                    return f"{f:.0f}%"
            except Exception:
                pass
            return s
        if isinstance(v, (int, float)):
            if 0 <= v <= 1.2:
                return f"{v*100:.0f}%"
            if 1.2 < v <= 100:
                return f"{v:.0f}%"
        return "tbd"

    def _to_int_or_tbd(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "tbd"
        if isinstance(v, str):
            s = v.strip()
            if not s or s.lower() == "tbd":
                return "tbd"
            if s.upper() == "NA":
                return "NA"
            try:
                return int(float(s.replace(",", ".")))
            except Exception:
                return s
        if isinstance(v, (int, float)):
            try:
                return int(v)
            except Exception:
                return "tbd"
        return "tbd"
    
    # Dictionnaire par d√©faut
    crit = {
        "pc":  {"nb": 764, "critere": "99%", "echelle": "moins de 1% de FN soit 99% de TN"},
        "lc":  {"nb": 764, "critere": "99%", "echelle": "moins de 1% de FN soit 99% de TN"},
        "c":   {"nb": 252, "critere": "97%", "echelle": "moins de 3% de FN soit 97% de TN"},
        "tbd": {"nb": "tbd", "critere": "tbd", "echelle": ""},
        "NA":  {"nb": "NA", "critere": "NA", "echelle": ""},
    }

    # Si l'Excel contient les colonnes, on enrichit/√©crase le mapping
    nb_col = None
    for c in ["nb test par fonction outil", "nb_tests", "nb test", "nb"]:
        if c in ref2.columns:
            nb_col = c
            break

    crit_col = None
    for c in ["crit√®re 1", "critere 1", "crit√®re", "critere", "critere %", "crit√®re %"]:
        if c in ref2.columns:
            crit_col = c
            break

    echelle_col = None
    for c in ["√©chelle", "echelle"]:
        if c in ref2.columns:
            echelle_col = c
            break

    if nb_col or crit_col or echelle_col:
        tmp_crit = ref2[["complexit√©"] + [c for c in [nb_col, crit_col, echelle_col] if c]].copy()
        tmp_crit["complexit√©"] = tmp_crit["complexit√©"].astype(str).str.strip()
        tmp_crit = tmp_crit.dropna(subset=["complexit√©"])
        tmp_crit = tmp_crit[tmp_crit["complexit√©"].str.lower().ne("nan")]

        for cx, grp in tmp_crit.groupby("complexit√©"):
            cx_norm = str(cx).strip()
            info = crit.get(cx_norm, {"nb": "tbd", "critere": "tbd", "echelle": ""}).copy()

            if nb_col and nb_col in grp.columns:
                val_nb = grp[nb_col].dropna().iloc[0] if grp[nb_col].dropna().shape[0] else None
                info["nb"] = _to_int_or_tbd(val_nb)

            if crit_col and crit_col in grp.columns:
                val_cr = grp[crit_col].dropna().iloc[0] if grp[crit_col].dropna().shape[0] else None
                info["critere"] = _to_percent_str(val_cr)

            if echelle_col and echelle_col in grp.columns:
                val_ec = grp[echelle_col].dropna().iloc[0] if grp[echelle_col].dropna().shape[0] else None
                info["echelle"] = "" if val_ec is None or (isinstance(val_ec, float) and pd.isna(val_ec)) else str(val_ec)

            crit[cx_norm] = info

    # =====================================================================
    # 1) Tableau ‚Äî Performance & quantit√© de tests par FS
    # =====================================================================
    st.subheader("1 ‚ÄîPerf & nb tests par FS")
    show_new = st.toggle("Afficher / masquer (1)", value=True, key="show_section_1")
    if show_new:
        # Quantit√© de tests : au niveau "test" (une ligne = un test)
        qte_tab = (
            data_f.groupby("fs_id")
            .agg(quantite_tests=("test_label", "count"))
            .reset_index()
        )

        # Performance : au niveau "fiche" (un fichier = un verdict)
        perf_base = data_calc_f.drop_duplicates(subset=["fs_id", "fiche_name"]).copy()
        perf_tab = (
            perf_base.groupby("fs_id")
            .agg(performance_score=("verdict_score", "mean"))
            .reset_index()
        )
        perf_tab["performance"] = perf_tab["performance_score"] * 100

        new_tab = (
            qte_tab.merge(perf_tab[["fs_id", "performance"]], on="fs_id", how="left")
            .rename(columns={"fs_id": "fsXX"})
        )

        # Tri naturel fs1, fs2, fs10...
        def _fs_key(s):
            m = re.search(r"(\d+)$", str(s))
            return int(m.group(1)) if m else 10**9

        new_tab = new_tab.sort_values("fsXX", key=lambda col: col.map(_fs_key))

        # Affichage avec % (mais on garde aussi une version num√©rique pour le graphique)
        new_tab["performance_num"] = pd.to_numeric(new_tab["performance"], errors="coerce")
        new_tab_display = new_tab[["fsXX", "performance_num", "quantite_tests"]].copy()
        new_tab_display["performance"] = new_tab_display["performance_num"].map(
            lambda x: "" if pd.isna(x) else f"{x:.1f}%"
        )
        new_tab_display = new_tab_display.rename(columns={"quantite_tests": "quantit√© de tests"})
        st.dataframe(new_tab_display[["fsXX", "performance", "quantit√© de tests"]], use_container_width=True)

        # Graphique double axe (comme la capture) : quantit√© (gauche) + performance % (droite)

        # Graphique double axe (comme la capture) : 2 barres par FS
        # - Quantit√© (bleu) sur l'axe gauche
        # - Performance (%) (rouge) sur l'axe droit
        # Graphique interactif (comme la capture) : 2 barres par FS
        # - Quantit√© (bleu clair) sur l'axe gauche
        # - Performance (%) (bleu fonc√©) sur l'axe droit
        # + Survol : d√©tail par classe documentaire
        fs_order = new_tab["fsXX"].astype(str).tolist()
        qte = pd.to_numeric(new_tab["quantite_tests"], errors="coerce").fillna(0).astype(int).tolist()
        perf = pd.to_numeric(new_tab["performance_num"], errors="coerce")
        perf_pct = perf.fillna(0).tolist()


        # D√©tails par classe documentaire
        tmp_detail = data_f.copy()
        tmp_detail["fs_id"] = tmp_detail["fs_id"].astype(str).str.strip()
        tmp_detail["classe_documentaire"] = tmp_detail.get("classe_documentaire", "").fillna("Inconnu").astype(str)

        tmp_perf_detail = data_calc_f.drop_duplicates(subset=["fs_id", "fiche_name"]).copy()
        tmp_perf_detail["classe_documentaire"] = tmp_perf_detail.get("classe_documentaire", "").fillna("Inconnu").astype(str)
        perf_by_class = (
            tmp_perf_detail.groupby(["fs_id", "classe_documentaire"])["verdict_score"]
            .mean()
            .reset_index()
        )
        perf_by_class["perf_pct"] = perf_by_class["verdict_score"] * 100
        qte_by_class = (
            tmp_detail.groupby(["fs_id", "classe_documentaire"])            .size()            .reset_index(name="nb")
        )

        def _hover_perf(fs: str, overall: float) -> str:
            sub = perf_by_class[perf_by_class["fs_id"] == fs].copy()
            sub = sub.sort_values("classe_documentaire")
            lines = [f"<b>{overall:.1f}%</b> (performance totale)"]
            for _, r in sub.iterrows():
                val = r["perf_pct"]
                if pd.notna(val):
                    lines.append(f"Classe doc {r['classe_documentaire']} : {val:.1f}%")
            return "<br>".join(lines)

        def _hover_qte(fs: str, total: int) -> str:
            sub = qte_by_class[qte_by_class["fs_id"] == fs].copy()
            sub = sub.sort_values("classe_documentaire")
            lines = [f"<b>{total} tests</b> (total)"]
            for _, r in sub.iterrows():
                lines.append(f"Classe doc {r['classe_documentaire']} : {int(r['nb'])}")
            return "<br>".join(lines)

        hover_perf = [_hover_perf(fs, float(p)) for fs, p in zip(fs_order, perf_pct)]
        hover_qte = [_hover_qte(fs, int(t)) for fs, t in zip(fs_order, qte)]

        fig_bar = go.Figure()
        fig_bar.add_bar(
            x=fs_order,
            y=qte,
            name="Quantit√© de tests",
            marker_color="#a7d2ff",
            hovertext=hover_qte,
            hoverinfo="text",
            offsetgroup="qte",
            alignmentgroup="fs",
        )
        fig_bar.add_bar(
            x=fs_order,
            y=perf_pct,
            name="Performance (%)",
            marker_color="#063970",
            yaxis="y2",
            hovertext=hover_perf,
            hoverinfo="text",
            offsetgroup="perf",
            alignmentgroup="fs",
        )

        fig_bar.update_layout(
            barmode="group",
            xaxis=dict(title="FS"),
            yaxis=dict(title="Quantit√© de tests", rangemode="tozero"),
            yaxis2=dict(title="Performance (%)", overlaying="y", side="right", range=[0, 100]),
            legend=dict(orientation="h", yanchor="bottom", y=1.15, xanchor="center", x=0.5),
            margin=dict(t=80, b=40, l=40, r=40),
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    # =====================================================================
    # 2) KPI globaux
    # =====================================================================
    # =====================================================================
    # 2) KPI globaux
    # =====================================================================
    st.subheader("2 ‚ÄîKPI globaux")

    # KPIs (sur les donn√©es filtr√©es)
    nb_tests_total = int(len(data_f))
    nb_docs = int(data_f["ref_coedm"].nunique()) if "ref_coedm" in data_f.columns else 0
    nb_fs = int(data_f["fs_id"].nunique()) if "fs_id" in data_f.columns else 0
    nb_testeurs = int(data_f["nom_testeur"].nunique()) if "nom_testeur" in data_f.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tests", nb_tests_total)
    c2.metric("Documents test√©s", nb_docs)
    c3.metric("FS test√©es", nb_fs)
    c4.metric("Testeurs", nb_testeurs)

    # -----------------------------------------------------------------
    # Remplacements NR vs TUF (NR √©crase TUF) + warning si NR < TUF
    # -----------------------------------------------------------------
    if isinstance(overrides_report, pd.DataFrame) and len(overrides_report):
        st.subheader("Remplacements NR vs TUF")
        st.info(
            f"{int(len(overrides_report))} cas : une fiche NR a remplac√© une fiche TUF (m√™me doc_id + m√™me FS)."
        )

        disp_all = overrides_report.copy()

        # Num√©riser pour d√©tecter NR < TUF
        for c in ["perf_TUF", "perf_NR", "delta_NR_minus_TUF"]:
            if c in disp_all.columns:
                disp_all[c] = pd.to_numeric(disp_all[c], errors="coerce")

        worse = disp_all.copy()
        if "delta_NR_minus_TUF" in worse.columns:
            worse = worse[worse["delta_NR_minus_TUF"] < 0]

        if len(worse):
            st.warning("‚ö†Ô∏è Attention : NR < TUF sur la performance moyenne pour certains remplacements.")

        disp_show = disp_all.copy()
        for c in ["perf_TUF", "perf_NR", "delta_NR_minus_TUF"]:
            if c in disp_show.columns:
                disp_show[c] = disp_show[c].map(lambda x: "" if pd.isna(x) else f"{x:.1f}%")

        st.dataframe(
            disp_show[
                ["doc_id", "fs_id", "fiche_TUF", "fiche_NR", "verdict_TUF", "verdict_NR", "perf_TUF", "perf_NR", "delta_NR_minus_TUF"]
            ],
            use_container_width=True,
        )

# ---------------------------------------------------------------------
    # Graphique unique : volume de tests + performance dans le temps
    # (axe X datetime => dates espac√©es proportionnellement)
    # ---------------------------------------------------------------------
    # ---------------------------------------------------------------------
    # Graphique : √©volution CUMUL√âE du nombre total de tests (courbe qui grimpe)
    # (axe X datetime => dates espac√©es proportionnellement)
    # ---------------------------------------------------------------------
    st.subheader("3 ‚Äî√âvolution cumul√©e tests")

    if "date_test" in data_f.columns:
        import matplotlib.dates as mdates

        tmp = data_f.copy()
        tmp["date_test"] = pd.to_datetime(tmp["date_test"], errors="coerce")
        tmp = tmp.dropna(subset=["date_test"])

        if len(tmp):
            # Agr√©gation par jour
            daily = (
                tmp.groupby(tmp["date_test"].dt.floor("D"))
                .agg(nb_tests=("date_test", "size"))
                .reset_index()
                .rename(columns={"date_test": "date"})
                .sort_values("date")
            )

            daily_cum = daily[["date", "nb_tests"]].copy()
            daily_cum["nb_tests_cum"] = daily_cum["nb_tests"].cumsum()

            fig2, ax2 = plt.subplots(figsize=(11, 4))
            ax2.plot(
                daily_cum["date"],
                daily_cum["nb_tests_cum"],
                marker="o",
                color="tab:blue",
                label="Total cumul√© de tests",
            )
            ax2.set_ylabel("Nombre total de tests")
            ax2.grid(True, axis="y", alpha=0.3)

            locator2 = mdates.AutoDateLocator()
            ax2.xaxis.set_major_locator(locator2)
            ax2.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator2))
            fig2.autofmt_xdate()

            ax2.legend(loc="upper center", bbox_to_anchor=(0.5, 1.18), ncol=1, frameon=False)
            st.pyplot(fig2, use_container_width=True)
        else:
            st.info("Aucune date exploitable dans la colonne 'date_test'.")
    else:
        st.info("Aucune colonne 'date_test' d√©tect√©e : impossible d'afficher l'√©volution des tests dans le temps.")

    st.subheader("4 ‚ÄîJustesse moyenne FS")
    show_tab2bis = st.toggle("Afficher / masquer", value=True, key="show_2bis")
    if show_tab2bis:
        tab2bis, fo_cols_2bis = build_tab2bis(data_calc_f, ref_base)
        tab2bis_display = tab2bis[["N¬∞ fs"] + fo_cols_2bis]
        styles_2bis = style_tab2bis(tab2bis_display, fo_cols_2bis)
        styler_2bis = tab2bis_display.style.apply(lambda _: styles_2bis, axis=None)
        st.dataframe(styler_2bis, use_container_width=True)

    # =====================================================================
    # 4) Tableau 3 (progression)
    # =====================================================================
    st.subheader("5 ‚ÄîProgression tests FS")
    show_tab3 = st.toggle("Afficher / masquer", value=True, key="show_3")
    if show_tab3:
        fs_testes = data_f["fs_id"].unique()
        tab3 = ref_base[ref_base["N¬∞ fs"].isin(fs_testes)].copy()

        tests_counts = data_f.groupby("fs_id").size().to_dict()

        tmp = data_f.copy()
        tmp["vcat"] = tmp["verdict_doc"].apply(norm_verdict)

        tver = tmp.groupby(["fs_id", "vcat"]).size().unstack(fill_value=0)
        tests_vdict = {fs: row.to_dict() for fs, row in tver.iterrows()}

        def convert_cell_progress(row, fo):
            val = row[fo]
            cx = row["complexit√©"]
            fs = row["N¬∞ fs"]
            if val != "x":
                return ""
            done = tests_counts.get(fs, 0)
            info = crit.get(cx, {})
            total = info.get("nb", "tbd")

            if isinstance(total, str) and total in ("tbd", "NA"):
                return f"{done}/{total}"

            if isinstance(total, (int, float)) and total > 0:
                pct = done / total
                return f"{done}/{int(total)} ({pct:.1%})"

            return str(done)

        for fo in fo_cols:
            tab3[fo] = tab3.apply(lambda row: convert_cell_progress(row, fo), axis=1)

        def ratio_cat(fs, cat):
            fs = str(fs)
            total = tests_counts.get(fs, 0)
            if total == 0:
                return ""
            n = tests_vdict.get(fs, {}).get(cat, 0)
            return f"{int(n)}/{int(total)}"

        tab3["Bon"] = tab3["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "bon"))
        tab3["Partiellement bon"] = tab3["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "partiel"))
        tab3["Mauvais"] = tab3["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "mauvais"))
        tab3["None"] = tab3["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "none"))

        tab3_display = tab3.drop(columns=["complexit√©"])
        st.dataframe(tab3_display, use_container_width=True)

    # =====================================================================
    # 5) Tableau 1 ‚Äî R√©f√©rentiel brut
    # =====================================================================
    st.subheader("6 ‚ÄîR√©f√©rentiel brut")
    show_tab1 = st.toggle("Afficher / masquer", value=False, key="show_1")
    if show_tab1:
        st.dataframe(ref_base)

    # =====================================================================
    # 6) Tableau 2 ‚Äî Crit√®res par fonctionnalit√© (SANS code couleur)
    # =====================================================================
    st.subheader("7 ‚ÄîCrit√®res de r√©ussite par fonctionnalit√©")
    show_tab2 = st.toggle("Afficher / masquer", value=False, key="show_2")
    if show_tab2:
        tab2 = ref_base.copy()

        def convert_cell_percent(row, fo):
            val = row[fo]
            cx = row["complexit√©"]
            if val != "x":
                return ""
            if pd.isna(cx) or cx in ("tbd", "NA"):
                return "tbd"
            info = crit.get(cx)
            if not info:
                return "tbd"
            return info.get("critere", "tbd")

        for fo in fo_cols:
            tab2[fo] = tab2.apply(lambda row: convert_cell_percent(row, fo), axis=1)

        tab2 = tab2.drop(columns=["complexit√©"])
        st.dataframe(tab2)

    # (Section KPI globaux d√©plac√©e en 2)



    # =====================================================================
    # 8) R√©ussite par classe doc et fonctionnalit√©s
    # =====================================================================
    st.subheader("8 ‚ÄîR√©ussite par classe documentaire et fonctionnalit√©")
    show_cd = st.toggle("Afficher / masquer r√©ussite par classe doc", value=False)
    if show_cd:
        if "classe_documentaire" in data_f.columns:
            tmp_cd = data_f.copy()
            tmp_cd["vcat"] = tmp_cd["verdict_doc"].apply(norm_verdict)

            tot_cd = (
                tmp_cd.groupby(["fs_id", "classe_documentaire"])
                .size()
                .rename("total_tests")
            )

            bon_cd = (
                tmp_cd[tmp_cd["vcat"] == "bon"]
                .groupby(["fs_id", "classe_documentaire"])
                .size()
                .rename("bon_tests")
            )

            cd_df = pd.concat([tot_cd, bon_cd], axis=1).fillna(0)
            cd_df["bon_tests"] = cd_df["bon_tests"].astype(int)
            cd_df["total_tests"] = cd_df["total_tests"].astype(int)

            def fmt_ratio(row):
                if row["total_tests"] == 0:
                    return ""
                return f"{row['bon_tests']}/{row['total_tests']}"

            cd_df["ratio"] = cd_df.apply(fmt_ratio, axis=1)

            pivot_cd = cd_df.reset_index().pivot(
                index="fs_id",
                columns="classe_documentaire",
                values="ratio"
            )

            st.dataframe(pivot_cd)
        else:
            st.info("Aucune classe documentaire trouv√©e dans les fiches.")

    # =====================================================================
    # 9) Taux de justesse par classe doc
    # =====================================================================
    st.subheader("9 ‚ÄîJustesse par classe doc")
    show_taux_cd = st.toggle("Afficher / masquer taux par classe doc", value=False)
    if show_taux_cd:
        if "classe_documentaire" in data_f.columns:
            classe_summary = (
                data_f.groupby("classe_documentaire")
                .agg(
                    nb_tests=("test_label", "count"),
                    taux_justes_moy=("taux_justes", "mean"),
                )
                .reset_index()
            )
            classe_summary["taux_justes_moy"] = classe_summary["taux_justes_moy"].round(2)
            st.dataframe(classe_summary)
        else:
            st.info("Aucune classe documentaire trouv√©e dans les fiches.")
            classe_summary = pd.DataFrame()

    # =====================================================================
    # 10) KPI par fonctionnalit√©
    # =====================================================================
    st.subheader("10 ‚ÄîKPI par fonctionnalit√©")
    show_kpi_fct = st.toggle("Afficher / masquer KPI par fonctionnalit√©", value=False)
    if show_kpi_fct:
        kpi_fct = (
            data_f.groupby("fs_id")
            .agg(
                nb_tests=("test_label", "count"),
                temps_humain_moy=("temps_humain_s", "mean"),
                temps_machine_moy=("temps_machine_s", "mean"),
                taux_justes_moy=("taux_justes", "mean"),
                nb_docs=("ref_coedm", lambda s: s.nunique()),
            )
            .reset_index()
        )

        tmp_kpi = data_f.copy()
        tmp_kpi["vcat"] = tmp_kpi["verdict_doc"].apply(norm_verdict)

        counts = (
            tmp_kpi.groupby(["fs_id", "vcat"])
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )

        for c in ["bon", "partiel", "mauvais", "none"]:
            if c not in counts.columns:
                counts[c] = 0

        counts["total_tests"] = counts[["bon", "partiel", "mauvais", "none"]].sum(axis=1)

        def ratio_str(n, total):
            if total == 0:
                return ""
            return f"{int(n)}/{int(total)}"

        counts["bon_ratio"] = counts.apply(lambda r: ratio_str(r.get("bon", 0), r["total_tests"]), axis=1)
        counts["partiel_ratio"] = counts.apply(lambda r: ratio_str(r.get("partiel", 0), r["total_tests"]), axis=1)
        counts["mauvais_ratio"] = counts.apply(lambda r: ratio_str(r.get("mauvais", 0), r["total_tests"]), axis=1)

        kpi_fct = kpi_fct.merge(
            counts[["fs_id", "bon_ratio", "partiel_ratio", "mauvais_ratio"]],
            on="fs_id",
            how="left",
        )

        kpi_fct = kpi_fct.rename(columns={
            "bon_ratio": "bon",
            "partiel_ratio": "partiellement bon",
            "mauvais_ratio": "mauvais",
        })

        st.dataframe(kpi_fct)

    # =====================================================================
    # 11) Commentaires additionnels
    # =====================================================================
    st.subheader("11 ‚ÄîCommentaires additionnels d√©tect√©s")
    show_comments = st.toggle("Afficher / masquer commentaires", value=False)
    if show_comments:
        comment_rows = (
            data_f[["ref_coedm", "verdict_doc", "commentaire_additionnel"]]
            .dropna(subset=["commentaire_additionnel"])
            .drop_duplicates()
        )

        if comment_rows.empty:
            st.info("Aucun commentaire dans les fiches.")
        else:
            options = []
            for idx, row in comment_rows.iterrows():
                ref = row["ref_coedm"]
                verdict = row["verdict_doc"] or "Non renseign√©"
                label = f"{ref} ‚Äî Verdict : {verdict}"
                options.append((label, idx))

            labels = [o[0] for o in options]
            choice_label = st.selectbox(
                "S√©lectionner un document pour voir le commentaire",
                labels,
            )

            chosen_idx = dict(options)[choice_label]
            chosen_row = comment_rows.loc[chosen_idx]

            st.info(
                f"üìÑ {chosen_row['ref_coedm']} ‚Äî Verdict : {chosen_row['verdict_doc'] or 'Non renseign√©'}\n\n"
                f"üìù Commentaire : {chosen_row['commentaire_additionnel']}"
            )

    # =====================================================================
    # 12) Donn√©es extraites + exclusions (‚ùå)
    # =====================================================================
    st.subheader("12 ‚ÄîDonn√©es extraites")
    show_data = st.toggle("Afficher / masquer donn√©es extraites", value=False)
    if show_data:

        # Affichage complet (non filtr√©) ‚Äî SANS doc_id / test_uid
        df_clean_all = data[clean_cols].copy()
        if "date_test" in df_clean_all.columns:
            df_clean_all["date_test"] = pd.to_datetime(df_clean_all["date_test"], errors="coerce").dt.date

        num_cols = df_clean_all.select_dtypes(include=["float", "int"]).columns
        df_clean_all[num_cols] = df_clean_all[num_cols].round(2)

        # Editor pour exclure des tests (ID interne via l'index, non affich√©)
        df_editor = data[["test_uid", "fs_id", "ref_coedm", "test_label", "verdict_doc"]].copy()
        df_editor = df_editor.set_index("test_uid")
        df_editor["‚ùå Exclure"] = df_editor.index.isin(st.session_state["excluded_test_uids"])

        st.caption("Cochez ‚ùå Exclure pour ignorer des lignes (cela mettra √† jour tous les tableaux).")
        edited = st.data_editor(
            df_editor[["fs_id", "ref_coedm", "test_label", "verdict_doc", "‚ùå Exclure"]],
            hide_index=True,
            column_config={
                "fs_id": st.column_config.TextColumn("FS", disabled=True),
                "ref_coedm": st.column_config.TextColumn("Document", disabled=True),
                "test_label": st.column_config.TextColumn("Test", disabled=True),
                "verdict_doc": st.column_config.TextColumn("Verdict", disabled=True),
                "‚ùå Exclure": st.column_config.CheckboxColumn("‚ùå Exclure"),
            },
            disabled=["fs_id", "ref_coedm", "test_label", "verdict_doc"],
            use_container_width=True,
        )

        # Mettre √† jour la session_state
        new_excluded = set(edited.index[edited["‚ùå Exclure"]].astype(int).tolist())
        st.session_state["excluded_test_uids"] = new_excluded

        # Affichage data filtr√©e, avec couleur verdict (SANS doc_id / test_uid)
        df_clean_f = data_f[clean_cols].copy()
        if "date_test" in df_clean_f.columns:
            df_clean_f["date_test"] = pd.to_datetime(df_clean_f["date_test"], errors="coerce").dt.date
        num_cols_f = df_clean_f.select_dtypes(include=["float", "int"]).columns
        df_clean_f[num_cols_f] = df_clean_f[num_cols_f].round(2)

        st.dataframe(df_clean_f.style.applymap(color_verdict, subset=["verdict_doc"]))
    st.subheader("13 ‚ÄîR√©sultats par type doc")
    show_type = st.toggle("Afficher / masquer r√©sultats par type de doc", value=False)
    if show_type:
        summary = (
            data_f.groupby("type_document")
            .agg(
                nb_tests=("test_label", "count"),
                temps_humain_moy=("temps_humain_s", "mean"),
                temps_machine_moy=("temps_machine_s", "mean"),
                taux_justes_moy=("taux_justes", "mean"),
                taux_fn_moy=("taux_fn", "mean"),
                taux_fp_moy=("taux_fp", "mean"),
            )
            .reset_index()
        )
        st.dataframe(summary)

    # =====================================================================
    # 14) Exports
    # =====================================================================
    st.subheader("14 ‚ÄîExport")
    # 2) Export suivi recettage (Excel) ‚Äî filtr√©
    recettage_df = data_f[
        ["fs_id", "verdict_doc", "classe_documentaire", "site", "ref_coedm"]
    ].copy()

    recettage_df["score"] = recettage_df["verdict_doc"].apply(verdict_to_score)
    recettage_df["Modification"] = recettage_df["ref_coedm"].apply(extract_modification)

    recettage_df = recettage_df[
        ["fs_id", "verdict_doc", "score", "classe_documentaire", "site", "Modification"]
    ]

    from io import BytesIO
    output_recettage = BytesIO()
    with pd.ExcelWriter(output_recettage, engine="xlsxwriter") as writer:
        recettage_df.to_excel(writer, sheet_name="Suivi_recettage", index=False)

    st.download_button(
        "Export suivi recettage (Excel)",
        data=output_recettage.getvalue(),
        file_name="suivi_recettage.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # 3) Excel multi-feuilles avec un peu de mise en forme
    # --- Reconstruire Tableau 2bis / Tableau 3 (filtr√©) ---
    tab2bis_raw, fo_cols_2bis = build_tab2bis(data_calc_f, ref_base)
    tab2bis_xlsx = tab2bis_raw[["N¬∞ fs"] + fo_cols_2bis]

    fs_testes = data_f["fs_id"].unique()
    tab3_x = ref_base[ref_base["N¬∞ fs"].isin(fs_testes)].copy()
    tests_counts = data_f.groupby("fs_id").size().to_dict()

    def convert_cell_progress_export(row, fo):
        val = row[fo]
        cx = row["complexit√©"]
        fs = row["N¬∞ fs"]
        if val != "x":
            return ""
        done = tests_counts.get(fs, 0)
        info = crit.get(cx, {})
        total = info.get("nb", "tbd")

        if isinstance(total, str) and total in ("tbd", "NA"):
            return f"{done}/{total}"

        if isinstance(total, (int, float)) and total > 0:
            pct = done / total
            return f"{done}/{int(total)} ({pct:.1%})"

        return str(done)

    for fo in fo_cols:
        tab3_x[fo] = tab3_x.apply(lambda row: convert_cell_progress_export(row, fo), axis=1)

    tmp_export = data_f.copy()
    tmp_export["vcat"] = tmp_export["verdict_doc"].apply(norm_verdict)
    tver = tmp_export.groupby(["fs_id", "vcat"]).size().unstack(fill_value=0)
    tests_vdict = {fs: row.to_dict() for fs, row in tver.iterrows()}

    def ratio_cat(fs, cat):
        fs = str(fs)
        total = tests_counts.get(fs, 0)
        if total == 0:
            return ""
        n = tests_vdict.get(fs, {}).get(cat, 0)
        return f"{int(n)}/{int(total)}"

    tab3_x["Bon"] = tab3_x["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "bon"))
    tab3_x["Partiellement bon"] = tab3_x["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "partiel"))
    tab3_x["Mauvais"] = tab3_x["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "mauvais"))
    tab3_x["None"] = tab3_x["N¬∞ fs"].map(lambda fs: ratio_cat(fs, "none"))

    tab3_display = tab3_x.drop(columns=["complexit√©"])

    # Donn√©es extraites filtr√©es
    clean_cols_export = clean_cols
    df_clean_export = data_f[clean_cols_export].copy()
    if "date_test" in df_clean_export.columns:
        df_clean_export["date_test"] = pd.to_datetime(df_clean_export["date_test"], errors="coerce").dt.date
    num_cols = df_clean_export.select_dtypes(include=["float", "int"]).columns
    df_clean_export[num_cols] = df_clean_export[num_cols].round(2)

    # KPI par fonctionnalit√© (filtr√©)
    kpi_fct_export = (
        data_f.groupby("fs_id")
        .agg(
            nb_tests=("test_label", "count"),
            temps_humain_moy=("temps_humain_s", "mean"),
            temps_machine_moy=("temps_machine_s", "mean"),
            taux_justes_moy=("taux_justes", "mean"),
            nb_docs=("ref_coedm", lambda s: s.nunique()),
        )
        .reset_index()
    )

    # R√©sultat par type doc (filtr√©)
    summary_export = (
        data_f.groupby("type_document")
        .agg(
            nb_tests=("test_label", "count"),
            temps_humain_moy=("temps_humain_s", "mean"),
            temps_machine_moy=("temps_machine_s", "mean"),
            taux_justes_moy=("taux_justes", "mean"),
            taux_fn_moy=("taux_fn", "mean"),
            taux_fp_moy=("taux_fp", "mean"),
        )
        .reset_index()
    )

    # Tableau perf/quantit√© par FS (filtr√©) ‚Äî pour l'export
    perfqte_export = (
        data_f.groupby('fs_id')
        .agg(performance=('taux_justes','mean'), quantite_tests=('test_label','count'))
        .reset_index()
        .rename(columns={'fs_id':'fs'})
    )
    perfqte_export['performance'] = pd.to_numeric(perfqte_export['performance'], errors='coerce')
    perfqte_export['performance_%'] = perfqte_export['performance'].map(lambda x: '' if pd.isna(x) else f"{x:.1f}%")

    output = BytesIO()

    # Titres (doivent √™tre identiques aux sous-titres Streamlit et compatibles Excel)
    TITLE_PERF_FS = "1 ‚ÄîPerf & nb tests par FS"
    TITLE_JUSTESSE_FS = "4 ‚ÄîJustesse moyenne FS"
    TITLE_PROG_FS = "5 ‚ÄîProgression tests FS"
    TITLE_REPL = "Remplacements NR vs TUF"
    TITLE_JUSTESSE_CD = "9 ‚ÄîJustesse par classe doc"
    TITLE_KPI_FCT = "10 ‚ÄîKPI par fonctionnalit√©"
    TITLE_DATA = "12 ‚ÄîDonn√©es extraites"
    TITLE_TYPE_DOC = "13 ‚ÄîR√©sultats par type doc"

    # 1) Perf / quantit√© par FS (filtr√©)
    perf_fs_export = (
        perfqte_export[["fs", "performance_%", "quantite_tests"]]
        .rename(columns={"fs": "fsXX", "performance_%": "performance", "quantite_tests": "quantit√© de tests"})
        .copy()
    )

    # 10) Justesse par classe doc (filtr√©)
    justesse_cd_export = pd.DataFrame()
    if "classe_documentaire" in data_f.columns:
        justesse_cd_export = (
            data_f.groupby("classe_documentaire")
            .agg(
                nb_tests=("test_label", "count"),
                taux_justes_moy=("taux_justes", "mean"),
            )
            .reset_index()
        )
        justesse_cd_export["taux_justes_moy"] = justesse_cd_export["taux_justes_moy"].round(2)

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        perf_fs_export.to_excel(writer, sheet_name=excel_sheet_name(TITLE_PERF_FS), index=False)
        tab2bis_xlsx.to_excel(writer, sheet_name=excel_sheet_name(TITLE_JUSTESSE_FS), index=False)
        tab3_display.to_excel(writer, sheet_name=excel_sheet_name(TITLE_PROG_FS), index=False)

        if isinstance(overrides_report, pd.DataFrame) and len(overrides_report):
            overrides_report.to_excel(writer, sheet_name=excel_sheet_name(TITLE_REPL), index=False)

        if isinstance(justesse_cd_export, pd.DataFrame) and len(justesse_cd_export):
            justesse_cd_export.to_excel(writer, sheet_name=excel_sheet_name(TITLE_JUSTESSE_CD), index=False)

        kpi_fct_export.to_excel(writer, sheet_name=excel_sheet_name(TITLE_KPI_FCT), index=False)
        summary_export.to_excel(writer, sheet_name=excel_sheet_name(TITLE_TYPE_DOC), index=False)
        df_clean_export.to_excel(writer, sheet_name=excel_sheet_name(TITLE_DATA), index=False)

        # Formats Excel (couleurs)
        workbook = writer.book

        fmt_red = workbook.add_format({"bg_color": "#fb3f35"})
        fmt_orange = workbook.add_format({"bg_color": "#F4B183"})
        fmt_yellow = workbook.add_format({"bg_color": "#FFF2CC"})
        fmt_light_green = workbook.add_format({"bg_color": "#C6EFCE"})
        fmt_dark_green = workbook.add_format({"bg_color": "#00B050", "font_color": "#FFFFFF"})

        fmt_verdict_bon = workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#006100", "bold": True})
        fmt_verdict_partiel = workbook.add_format({"bg_color": "#FFE699", "font_color": "#7F6000", "bold": True})
        fmt_verdict_mauvais = workbook.add_format({"bg_color": "#fb3f35", "font_color": "#9C0006", "bold": True})

        # Couleurs % dans "Justesse moyenne FS"
        ws2 = writer.sheets[excel_sheet_name(TITLE_JUSTESSE_FS)]
        n_rows_2 = len(tab2bis_xlsx)

        def _simple_pct(val):
            if not isinstance(val, str):
                return None
            s = val.strip().replace("%", "").replace(",", ".")
            if not s:
                return None
            try:
                return float(s)
            except Exception:
                return None

        for r in range(n_rows_2):
            for fo in fo_cols_2bis:
                c_idx = tab2bis_xlsx.columns.get_loc(fo)
                val = tab2bis_xlsx.iloc[r, c_idx]
                pct = _simple_pct(val)
                if pct is None:
                    continue
                if pct < 25:
                    fmt = fmt_red
                elif pct < 50:
                    fmt = fmt_orange
                elif pct < 75:
                    fmt = fmt_yellow
                elif pct < 90:
                    fmt = fmt_light_green
                else:
                    fmt = fmt_dark_green
                ws2.write(r + 1, c_idx, val, fmt)

        # Couleurs verdict dans "Donn√©es extraites"
        wsD = writer.sheets[excel_sheet_name(TITLE_DATA)]
        n_rows_D = len(df_clean_export)
        verdict_col_idx = df_clean_export.columns.get_loc("verdict_doc")

        for r in range(n_rows_D):
            val = df_clean_export.iloc[r, verdict_col_idx]
            if not isinstance(val, str):
                continue
            v = val.strip().lower()
            if v == "bon":
                fmt = fmt_verdict_bon
            elif "partiellement" in v:
                fmt = fmt_verdict_partiel
            elif "mauvais" in v:
                fmt = fmt_verdict_mauvais
            else:
                continue
            wsD.write(r + 1, verdict_col_idx, val, fmt)


    st.download_button(
        "Exporter les tableaux (Excel)",
        data=output.getvalue(),
        file_name="IA4Doc_tableaux_complets_filtre.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
